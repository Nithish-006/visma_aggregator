"""Per-project Excel export for the Project Summary page and registry pop-up.

Both /api/project-summary/export (single selected project) and
/api/projects/<id>/export (registry pop-up) build one workbook per project,
structured like the registry detail pop-up. openpyxl is imported lazily inside
the builders. The legacy multi-project/aggregated report was removed.
"""

import io
import re

import pandas as pd

from flask import request, jsonify, send_file

from config import VALID_BANK_CODES, now_ist
from extensions import db_manager
import salary_api
from helpers.bankdata import get_bank_df
from helpers.dataframe import filter_by_date_range
from helpers.projects import parse_project_selection
from helpers.project_finance import compute_project_finance


def export_project_summary():
    """Export the selected project's report as the pop-up-structured workbook.

    The Project Summary page only shows its Export button once a single project
    is selected, so an export always targets exactly one project. This resolves
    the canonical selection ("<id> - NAME") to a project id and delegates to
    export_single_project_summary(). There is no multi-project / aggregated
    export — that legacy multi-tab report was removed.
    """
    project = request.args.get('project', None)
    single_id = None
    if project:
        sel_ids, sel_stems = parse_project_selection(project)
        if len(sel_ids) == 1 and not sel_stems:
            single_id = sel_ids[0]
        elif not sel_ids and len(sel_stems) == 1:
            try:
                match = db_manager.find_project_by_stem(sel_stems[0])
            except Exception:
                match = None
            if match:
                single_id = match['id']
    if single_id is None:
        return jsonify({
            'error': 'select_one_project',
            'message': 'Select a single project to export its summary.',
        }), 400
    return export_single_project_summary(single_id)


# ════════════════════════════════════════════════════════════════════════════
# PER-PROJECT EXPORT
#
# One project = one workbook, structured exactly like the registry detail
# pop-up: PO Value → Client Payments → Expenses → Purchase Bills → Sales Bills
# → Labour Payments → Consolidated Summary. The Expenses sheet consolidates the
# category breakdown (with a bar chart), vendor + bank breakdowns, and a
# filterable transactions table.
# ════════════════════════════════════════════════════════════════════════════

# Category buckets — mirror blueprints.projects.api_project_insights so the
# Consolidated Summary's spend composition matches the pop-up's headline.
_LABOUR_CATS = {'LABOUR PAYMENT', 'LABOR PAYMENT', 'LABOUR', 'LABOR'}
_OTHER_EXCLUDE_CATS = {'MATERIAL PURCHASE', 'AMOUNT RECEIVED', 'SALARY AC',
                       'BANK CHARGES', 'DUTIES & TAX'}


def _first_present(*values):
    """First value that isn't None — zero and 0.00 are answers, not absences.

    A plain `a or b` reaches for the fallback whenever the leading value is a
    legitimate zero, which for PO figures means silently reporting a different
    contract than the one asked for.
    """
    for v in values:
        if v is not None:
            return v
    return None


def _make_styles():
    """Build the shared openpyxl style set (lazy import, as elsewhere)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        'header_font': Font(name='Calibri', bold=True, color='FFFFFF', size=11),
        'header_fill': PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid'),
        'title_font': Font(name='Calibri', bold=True, size=14, color='1A1A2E'),
        'subtitle_font': Font(name='Calibri', bold=True, size=11, color='4A4A68'),
        'currency_fmt': '#,##0.00',
        'pct_fmt': '0.0%',
        'thin_border': Border(bottom=Side(style='thin', color='E5E7EB')),
        'income_font': Font(name='Calibri', color='059669', bold=True),
        'expense_font': Font(name='Calibri', color='DC2626', bold=True),
        'section_bold': Font(name='Calibri', bold=True, size=11),
        'green_amount': Font(name='Calibri', bold=True, color='006100'),
        'red_amount': Font(name='Calibri', bold=True, color='DC2626'),
        'blue_amount': Font(name='Calibri', bold=True, color='2563EB'),
        'green_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'block_bg': PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid'),
        'project_name_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'project_name_font': Font(name='Calibri', bold=True, size=14, color='FFFFFF'),
        'separator_fill': PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid'),
        # Bank-coloured fonts, mirroring the pop-up's Axis-red / KVB-green badges.
        'bank_axis_font': Font(name='Calibri', bold=True, color='C0392B'),
        'bank_kvb_font': Font(name='Calibri', bold=True, color='1E8449'),
    }


def _write_project_bills_sheet(ws, bills, project_label, sheet_title,
                               party_label, party_key, gstin_key, total_font, st):
    """Project-grouped bills sheet for a SINGLE project (one project block).

    A trimmed sibling of export_project_summary's write_bills_sheet: same
    auditor layout (project header → per-bill line items → bill totals →
    project total), but for just this project's bills.
    """
    from openpyxl.styles import Font, Alignment, PatternFill

    title_font = st['title_font']
    header_font = st['header_font']
    header_fill = st['header_fill']
    currency_fmt = st['currency_fmt']
    thin_border = st['thin_border']
    green_fill = st['green_fill']
    block_bg = st['block_bg']
    project_name_fill = st['project_name_fill']
    project_name_font = st['project_name_font']

    BILL_COLS = 15
    bill_header_labels = [
        'SL.NO', party_label, 'GSTIN', 'INVOICE #', 'DATE',
        'DESCRIPTION', 'HSN/SAC', 'QTY', 'UOM', 'RATE',
        'TAXABLE AMT', 'CGST', 'SGST', 'IGST', 'TOTAL'
    ]

    ws.cell(row=1, column=1, value=sheet_title).font = title_font
    cr = 3
    block_start = cr

    # ── PROJECT HEADER (blue fill, white text) ──
    for c in range(1, BILL_COLS + 1):
        ws.cell(row=cr, column=c).fill = project_name_fill
    ws.cell(row=cr, column=1, value=f'PROJECT :  {project_label.upper()}').font = project_name_font
    cr += 2

    # ── COLUMN HEADERS ──
    for ci, lbl in enumerate(bill_header_labels, 1):
        cell = ws.cell(row=cr, column=ci, value=lbl)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    cr += 1

    grand_taxable = grand_cgst = grand_sgst = grand_igst = grand_total = 0
    bill_serial = 0

    for bill in bills:
        bill_serial += 1
        line_items = bill.get('line_items', [])
        b_taxable = float(bill.get('subtotal', 0) or 0)
        b_cgst = float(bill.get('total_cgst', 0) or 0)
        b_sgst = float(bill.get('total_sgst', 0) or 0)
        b_igst = float(bill.get('total_igst', 0) or 0)
        b_total = float(bill.get('total_amount', 0) or 0)

        if line_items:
            for li_idx, item in enumerate(line_items):
                if li_idx == 0:
                    ws.cell(row=cr, column=1, value=bill_serial)
                    ws.cell(row=cr, column=2, value=bill.get(party_key, ''))
                    ws.cell(row=cr, column=3, value=bill.get(gstin_key, ''))
                    ws.cell(row=cr, column=4, value=bill.get('invoice_number', ''))
                    ws.cell(row=cr, column=5, value=bill.get('invoice_date', ''))
                ws.cell(row=cr, column=6, value=item.get('description', ''))
                ws.cell(row=cr, column=7, value=item.get('hsn_sac_code', ''))
                qty = item.get('quantity', 0)
                if qty:
                    ws.cell(row=cr, column=8, value=qty).number_format = '#,##0.00'
                ws.cell(row=cr, column=9, value=item.get('uom', ''))
                rate = item.get('rate_per_unit', 0)
                if rate:
                    ws.cell(row=cr, column=10, value=rate).number_format = currency_fmt
                taxable = item.get('taxable_value', 0)
                if taxable:
                    ws.cell(row=cr, column=11, value=taxable).number_format = currency_fmt
                if item.get('cgst_amount', 0):
                    ws.cell(row=cr, column=12, value=item['cgst_amount']).number_format = currency_fmt
                if item.get('sgst_amount', 0):
                    ws.cell(row=cr, column=13, value=item['sgst_amount']).number_format = currency_fmt
                if item.get('igst_amount', 0):
                    ws.cell(row=cr, column=14, value=item['igst_amount']).number_format = currency_fmt
                cr += 1
        else:
            ws.cell(row=cr, column=1, value=bill_serial)
            ws.cell(row=cr, column=2, value=bill.get(party_key, ''))
            ws.cell(row=cr, column=3, value=bill.get(gstin_key, ''))
            ws.cell(row=cr, column=4, value=bill.get('invoice_number', ''))
            ws.cell(row=cr, column=5, value=bill.get('invoice_date', ''))
            cr += 1

        # ── Bill Total row ──
        ws.cell(row=cr, column=6, value='Bill Total').font = Font(bold=True)
        for col, val in ((11, b_taxable), (12, b_cgst), (13, b_sgst), (14, b_igst), (15, b_total)):
            cell = ws.cell(row=cr, column=col, value=val)
            cell.font = Font(bold=True)
            cell.number_format = currency_fmt
        for c in range(1, BILL_COLS + 1):
            ws.cell(row=cr, column=c).border = thin_border
        cr += 1

        grand_taxable += b_taxable
        grand_cgst += b_cgst
        grand_sgst += b_sgst
        grand_igst += b_igst
        grand_total += b_total

    # ── PROJECT / GRAND TOTAL (green fill) ──
    for c in range(1, BILL_COLS + 1):
        ws.cell(row=cr, column=c).fill = green_fill
    ws.cell(row=cr, column=1, value=f'PROJECT TOTAL — {project_label.upper()}').font = Font(bold=True)
    for col, val in ((11, grand_taxable), (12, grand_cgst), (13, grand_sgst),
                     (14, grand_igst), (15, grand_total)):
        cell = ws.cell(row=cr, column=col, value=val)
        cell.font = total_font
        cell.number_format = currency_fmt

    # ── Mild-yellow background over the block's content rows (line items + bill
    # totals); the blue project header / column headers and the green project
    # total keep their own fills, matching the prior auditor theme. ──
    blank = PatternFill(fill_type=None)
    for rr in range(block_start, cr + 1):
        for c in range(1, BILL_COLS + 1):
            cell = ws.cell(row=rr, column=c)
            if cell.fill == blank or cell.fill == PatternFill():
                cell.fill = block_bg

    col_widths = {'A': 8, 'B': 28, 'C': 18, 'D': 18, 'E': 14, 'F': 35, 'G': 12,
                  'H': 10, 'I': 8, 'J': 12, 'K': 15, 'L': 12, 'M': 12, 'N': 12, 'O': 15}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


def export_single_project_summary(project_id):
    """Per-project Excel export, structured like the registry detail pop-up.

    Sheets, in pop-up order:
      1. PO Value
      2. Client Payments (KVB bank credits + cash ledger)
      3. Expenses
      4. Purchase Bills
      5. Sales Bills
      6. Labour Payments
      7. Consolidated Summary
    Then the analytical extras, scoped to this project:
      Expense by Category, Vendor Breakdown, Axis Txns, KVB Txns.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    db_manager.ensure_projects_table()
    project = db_manager.get_project(project_id)
    if not project:
        return jsonify({'error': 'not_found'}), 404

    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)

    if start_date and end_date:
        date_label = f"Period: {start_date} to {end_date}"
    elif start_date:
        date_label = f"Period: from {start_date}"
    elif end_date:
        date_label = f"Period: up to {end_date}"
    else:
        date_label = ''

    display = project.get('display') or f"{project_id} - {project.get('stem_name', '')}"
    stem_name = project.get('stem_name', '') or display
    # Full PO gist (terms, tax split, payment terms, scope line items) so the
    # PO Value sheet mirrors everything we persist — nothing dropped to a headline.
    try:
        po_gist = db_manager.get_project_po(project_id) or {}
    except Exception as e:
        print(f"[!] Error fetching PO gist for project export: {e}")
        po_gist = {}
    # The joined column already includes agreed variations, so it wins outright.
    # Testing it for falsiness rather than None would hand the baseline back to a
    # contract whose variations happen to cancel its PO out to zero.
    po_value = float(_first_present(project.get('po_total_value'),
                                    po_gist.get('total_value')) or 0)
    # None here means no PO at all, which is not the same as a contract varied
    # down to zero — the receivable rule below turns on that distinction.
    has_po = _first_present(project.get('po_total_value'), po_gist.get('total_value')) is not None
    try:
        variations = db_manager.list_po_variations(project_id)
    except Exception as e:
        print(f"[!] Error fetching PO variations for project export: {e}")
        variations = []
    var_total = sum(float(v.get('total_amount') or 0) for v in variations)

    st = _make_styles()
    currency_fmt = st['currency_fmt']
    pct_fmt = st['pct_fmt']

    # ── Gather this project's bank transactions across banks ──
    def project_mask(df):
        col = 'Project' if 'Project' in df.columns else 'project'
        if col not in df.columns:
            return pd.Series(False, index=df.index)
        s = df[col].astype(str).str.strip()
        return s.str.match(rf'^{project_id}\s*-')

    combined_rows = []
    for bc in VALID_BANK_CODES:
        df = get_bank_df(bc).copy()
        if df.empty:
            continue
        df = filter_by_date_range(df, start_date, end_date)
        df = df[project_mask(df)]
        if df.empty:
            continue
        df['bank'] = bc
        combined_rows.append(df)
    combined = pd.concat(combined_rows, ignore_index=True) if combined_rows else pd.DataFrame()

    # ── Client payments: KVB credits + cash ledger ──
    kvb_df = combined[combined['bank'] == 'kvb'] if not combined.empty else pd.DataFrame()
    if not kvb_df.empty and 'CR Amount' in kvb_df.columns:
        bank_credits = kvb_df[kvb_df['CR Amount'] > 0].sort_values('date', ascending=False)
    else:
        bank_credits = pd.DataFrame()
    received_bank = float(bank_credits['CR Amount'].sum()) if not bank_credits.empty else 0.0
    cash_payments = db_manager.list_cash_payments(project_id)
    received_cash = sum(float(c.get('amount') or 0) for c in cash_payments)
    received_total = received_bank + received_cash

    # ── Expenses: debit transactions, all banks ──
    expense_df = combined[combined['DR Amount'] > 0] if not combined.empty else pd.DataFrame()
    v_col = 'Client/Vendor' if (not combined.empty and 'Client/Vendor' in combined.columns) else 'client_vendor'

    # ── Bills with line items, filtered to this project ──
    def _belongs(bill):
        proj = str(bill.get('project', '') or '').strip()
        return bool(re.match(rf'^{project_id}\s*-', proj))

    try:
        purchase_bills = [b for b in db_manager.get_bills_with_line_items_for_export(
            start_date=start_date, end_date=end_date) if _belongs(b)]
    except Exception as e:
        print(f"[!] Error fetching purchase bills for project export: {e}")
        purchase_bills = []
    try:
        sales_bills = [b for b in db_manager.get_sales_bills_with_line_items_for_export(
            start_date=start_date, end_date=end_date) if _belongs(b)]
    except Exception as e:
        print(f"[!] Error fetching sales bills for project export: {e}")
        sales_bills = []

    material_total = sum(float(b.get('total_amount') or 0) for b in purchase_bills)

    # ── Labour from the salary API ──
    try:
        labour_summary = salary_api.get_labour_summary_for_project(
            project_id, project_display=display, start_date=start_date, end_date=end_date)
    except Exception as e:
        print(f"[!] Labour summary error (project {project_id}): {e}")
        labour_summary = {'available': False, 'monthly': [], 'total_cost': 0.0}
    labour_total = float(labour_summary.get('total_cost') or 0)
    try:
        labour_monthly_detail = salary_api.get_monthly_labour_summary(
            start_date=start_date, end_date=end_date, project=display)
    except Exception as e:
        print(f"[!] Labour monthly detail error (project {project_id}): {e}")
        labour_monthly_detail = []

    # ── Other (non-material, non-labour, non-internal) expense ──
    other_expense_total = 0.0
    if not expense_df.empty and 'Category' in expense_df.columns:
        up = expense_df['Category'].astype(str).str.upper().str.strip()
        keep = ~(up.isin(_OTHER_EXCLUDE_CATS) | up.isin(_LABOUR_CATS))
        other_expense_total = float(expense_df[keep]['DR Amount'].sum())

    # ── Value ladder + GST position ──
    # Shared with the registry detail pop-up via helpers/project_finance so the
    # sheet and the pop-up can't drift; this workbook is that pop-up in Excel.
    def _bill_sums(bills):
        return {
            'taxable': sum(float(b.get('subtotal') or 0) for b in bills),
            'gst': sum(float(b.get('total_cgst') or 0) + float(b.get('total_sgst') or 0)
                       + float(b.get('total_igst') or 0) for b in bills),
            'total': sum(float(b.get('total_amount') or 0) for b in bills),
        }

    sales = _bill_sums(sales_bills)
    purchase = _bill_sums(purchase_bills)

    # These lists are date-filtered. Whether the project HAS sales bills is not
    # a question about the period, so ask it unfiltered — otherwise a range that
    # excludes every sales bill would silently fall back to the full PO value
    # and report it against period-scoped costs as a fabricated profit.
    try:
        _, all_sales_summary = db_manager.get_bills_for_canonical_project(
            project_id, kind='sales', limit=1)
        has_sales_bills = all_sales_summary['total_amount'] > 0
    except Exception as e:
        print(f"[!] Error checking sales bills for project {project_id}: {e}")
        has_sales_bills = sales['total'] > 0

    fin = compute_project_finance(
        sales=sales,
        purchase=purchase,
        # The joined columns lead, because they carry the agreed variations
        # folded in; po_gist holds only the baseline as extracted, so reading it
        # first would export the contract at its signed value and quietly drop
        # every change since. It stays as the fallback for a project whose gist
        # row exists but never made it into the join.
        po={'taxable': _first_present(project.get('po_taxable_value'), po_gist.get('taxable_value')),
            'gst': _first_present(project.get('po_total_tax'), po_gist.get('total_tax')),
            'total': po_value},
        received_total=received_total,
        other_expense_total=other_expense_total,
        labour_total=labour_total,
        overhead=project.get('overhead'),
        has_sales_bills=has_sales_bills,
        has_po=has_po,
    )
    value_basic = fin['value']['basic']
    value_gst = fin['value']['gst']
    value_total = fin['value']['total']
    sales_taxable, sales_gst = sales['taxable'], sales['gst']
    purchase_taxable, purchase_gst = purchase['taxable'], purchase['gst']
    gst_extra = fin['gst']['extra']
    gst_extra_cost = fin['gst']['extra_cost']
    overhead = fin['overhead']
    spend_total = fin['spend_total']
    profit = fin['profit']
    # What the client still owes against the project's value.
    balance = fin['receivable']
    # Sheet 1 reconciles against the PO specifically (it prints PO Value and a
    # "% Received" computed from it), so it keeps its own PO-based balance
    # rather than the value-ladder one.
    po_balance = po_value - received_total

    wb = Workbook()
    wb.remove(wb.active)

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = st['header_font']
            cell.fill = st['header_fill']
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Banner fills carry long labels in column 1 (project header, project total,
    # column-header band) that Excel spills over empty neighbours — they must not
    # dictate column A's width.
    _BANNER_FILLS = {'4472C4', 'C6EFCE', '2563EB', '2F2F2F'}

    def _is_col1_banner(cell):
        f = cell.fill
        if f is not None and getattr(f, 'patternType', None) == 'solid':
            rgb = (f.fgColor.rgb or '')[-6:].upper()
            return rgb in _BANNER_FILLS
        return False

    def auto_width(ws, cap=50):
        """Expand columns to fit content so the sheet opens fully readable.

        The title row and column-1 banner cells (which spill in Excel) are
        excluded so they don't blow out column A, and any preset column width is
        treated as a floor — auto_width only ever widens, never narrows.
        """
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.row == 1:                       # decorative title row
                    continue
                if cell.column == 1 and _is_col1_banner(cell):
                    continue
                try:
                    val = str(cell.value) if cell.value is not None else ''
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            existing = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(existing, min(max_len + 3, cap))

    def kv_row(ws, row, label, value, *, fmt=None, font=None, bold_label=True):
        lcell = ws.cell(row=row, column=1, value=label)
        if bold_label:
            lcell.font = Font(bold=True)
        vcell = ws.cell(row=row, column=2, value=value)
        if fmt:
            vcell.number_format = fmt
        if font:
            vcell.font = font

    def bank_font(bc):
        return st['bank_axis_font'] if bc == 'axis' else st['bank_kvb_font'] if bc == 'kvb' else None

    # ─────────────────────────────────────────────── SHEET 1: PO Value ──
    # Comprehensive PO breakdown: every term we persist (number/date/client,
    # the taxable + tax + total split, amount-in-words, payment terms, the PO
    # document) followed by the full scope line-item table, then the received /
    # balance roll-up. Nothing is collapsed to a single headline value.
    ws = wb.create_sheet('PO Value')
    ws.cell(row=1, column=1, value=f'PO Value — {display}').font = st['title_font']
    pct = (received_total / po_value) if po_value > 0 else 0
    # This sheet reconciles against the PO alone, so its balance and % are both
    # PO-based. The value-ladder balance (which may be sales-bill sourced) is
    # reported on the Consolidated Summary instead — mixing the two here would
    # print a "Balance Due" that contradicts the "% Received" beside it.
    bal_label = 'Excess Received (vs PO)' if po_balance < -0.5 else 'Balance Due (vs PO)'

    r = 3
    ws.cell(row=r, column=1, value='PO TERMS').font = st['subtitle_font']; r += 1
    kv_row(ws, r, 'PO Number', po_gist.get('po_number') or project.get('po_number') or '—'); r += 1
    kv_row(ws, r, 'PO Date', po_gist.get('po_date') or '—'); r += 1
    kv_row(ws, r, 'Client', po_gist.get('client_name') or '—'); r += 1
    kv_row(ws, r, 'Currency', po_gist.get('currency') or 'INR'); r += 1
    kv_row(ws, r, 'PO Document', project.get('po_filename') or po_gist.get('source_filename') or '—'); r += 1
    if po_gist.get('extraction_status'):
        kv_row(ws, r, 'Extraction Status', str(po_gist.get('extraction_status'))); r += 1
    r += 1

    # ── Value split (taxable → tax → total) ──
    # Baseline figures, matching the scope line items printed below and the PO
    # document itself. Where variations exist they are shown as their own lines
    # rather than folded silently into these: po_value carries them, so printing
    # a variation-inclusive total over a baseline-only split would give the
    # client a breakdown that contradicts its own arithmetic with nothing on the
    # sheet to explain the gap.
    ws.cell(row=r, column=1, value='VALUE BREAKDOWN').font = st['subtitle_font']; r += 1
    base_taxable = float(po_gist.get('taxable_value') or 0)
    base_tax = float(po_gist.get('total_tax') or 0)
    kv_row(ws, r, 'Taxable Value', base_taxable, fmt=currency_fmt); r += 1
    kv_row(ws, r, 'Total Tax', base_tax, fmt=currency_fmt); r += 1
    if variations:
        kv_row(ws, r, 'PO Value (as extracted)', base_taxable + base_tax, fmt=currency_fmt); r += 1
        kv_row(ws, r, f'Variations ({len(variations)})', var_total, fmt=currency_fmt); r += 1
    kv_row(ws, r, 'PO Value (Total)', po_value, fmt=currency_fmt, font=st['blue_amount']); r += 1
    if po_gist.get('amount_in_words'):
        kv_row(ws, r, 'Amount in Words', str(po_gist.get('amount_in_words'))); r += 1
    if po_gist.get('payment_terms'):
        kv_row(ws, r, 'Payment Terms', str(po_gist.get('payment_terms'))); r += 1
    r += 1

    # ── Scope line items (description / qty / unit / rate / amount) ──
    line_items = po_gist.get('line_items') or []
    if line_items:
        ws.cell(row=r, column=1, value=f'SCOPE LINE ITEMS ({len(line_items)})').font = st['subtitle_font']; r += 1
        li_headers = ['SL.NO', 'Description', 'Qty', 'Unit', 'Rate', 'Amount']
        for ci, h in enumerate(li_headers, 1):
            ws.cell(row=r, column=ci, value=h)
        style_header_row(ws, r, len(li_headers))
        r += 1
        li_total = 0.0
        for idx, it in enumerate(line_items, 1):
            ws.cell(row=r, column=1, value=idx)
            ws.cell(row=r, column=2, value=str(it.get('description') or '—'))
            qv = it.get('quantity')
            if qv not in (None, '', 0):
                ws.cell(row=r, column=3, value=float(qv)).number_format = '#,##0.###'
            ws.cell(row=r, column=4, value=str(it.get('unit') or '—'))
            rv = it.get('rate')
            if rv not in (None, '', 0):
                ws.cell(row=r, column=5, value=float(rv)).number_format = currency_fmt
            av = it.get('amount')
            if av not in (None, '', 0):
                amt = float(av)
                ws.cell(row=r, column=6, value=amt).number_format = currency_fmt
                li_total += amt
            r += 1
        ws.cell(row=r, column=2, value='Line Items Total').font = Font(bold=True)
        tc = ws.cell(row=r, column=6, value=li_total)
        tc.font = Font(bold=True)
        tc.number_format = currency_fmt
        r += 2
    else:
        scope_n = po_gist.get('line_item_count')
        if scope_n:
            kv_row(ws, r, 'Scope Items', int(scope_n)); r += 2

    # ── Variations: the agreed changes between the PO and the contract ──
    # The scope items above are the PO as signed; these are what moved since.
    # Without them the sheet shows a total the client can't reconcile to the
    # document, so they are listed in full rather than netted into a figure.
    if variations:
        ws.cell(row=r, column=1, value=f'VARIATIONS ({len(variations)})').font = st['subtitle_font']; r += 1
        v_headers = ['SL.NO', 'Change', 'Weight', 'Unit', 'Rate', 'Basic', 'GST', 'Total']
        for ci, h in enumerate(v_headers, 1):
            ws.cell(row=r, column=ci, value=h)
        style_header_row(ws, r, len(v_headers))
        r += 1
        for idx, v in enumerate(variations, 1):
            ws.cell(row=r, column=1, value=idx)
            ws.cell(row=r, column=2, value=str(v.get('description') or '—'))
            ws.cell(row=r, column=3, value=float(v.get('quantity') or 0)).number_format = '#,##0.###'
            ws.cell(row=r, column=4, value=str(v.get('unit') or '—'))
            ws.cell(row=r, column=5, value=float(v.get('rate') or 0)).number_format = currency_fmt
            for col, key in ((6, 'basic_amount'), (7, 'tax_amount'), (8, 'total_amount')):
                ws.cell(row=r, column=col, value=float(v.get(key) or 0)).number_format = currency_fmt
            r += 1
        ws.cell(row=r, column=2, value='Net Change').font = Font(bold=True)
        nc = ws.cell(row=r, column=8, value=var_total)
        nc.font = Font(bold=True)
        nc.number_format = currency_fmt
        r += 2

    # ── Received / balance roll-up ──
    ws.cell(row=r, column=1, value='RECEIVED vs PO').font = st['subtitle_font']; r += 1
    kv_row(ws, r, 'Received — Bank (KVB)', received_bank, fmt=currency_fmt, font=st['income_font']); r += 1
    kv_row(ws, r, 'Received — Cash', received_cash, fmt=currency_fmt, font=st['income_font']); r += 1
    kv_row(ws, r, 'Total Received', received_total, fmt=currency_fmt, font=st['income_font']); r += 1
    kv_row(ws, r, bal_label, abs(po_balance), fmt=currency_fmt,
           font=(st['red_amount'] if po_balance > 0.5 else st['green_amount'])); r += 1
    kv_row(ws, r, '% Received', pct, fmt=pct_fmt); r += 1

    # ──────────────────────────────────────── SHEET 2: Client Payments ──
    ws = wb.create_sheet('Client Payments')
    ws.cell(row=1, column=1, value=f'Client Payments — {display}').font = st['title_font']
    # Chips row
    kv_row(ws, 3, 'Bank (KVB)', received_bank, fmt=currency_fmt, font=st['income_font'])
    ws.cell(row=3, column=3, value=f'{len(bank_credits)} credit(s)').font = Font(italic=True, color='6B7280')
    kv_row(ws, 4, 'Cash', received_cash, fmt=currency_fmt, font=st['income_font'])
    ws.cell(row=4, column=3, value=f'{len(cash_payments)} entr{"y" if len(cash_payments) == 1 else "ies"}').font = Font(italic=True, color='6B7280')
    kv_row(ws, 5, 'Total Received', received_total, fmt=currency_fmt, font=st['income_font'])
    # Combined history table
    headers = ['Date', 'Source', 'Particulars', 'Vendor', 'Amount']
    hr = 7
    for ci, h in enumerate(headers, 1):
        ws.cell(row=hr, column=ci, value=h)
    style_header_row(ws, hr, len(headers))
    pay_entries = []
    for _, row in bank_credits.iterrows():
        desc = ''
        for c in ('Description', 'Transaction Description', 'transaction_description'):
            if c in row and pd.notna(row.get(c)) and str(row.get(c)).strip():
                desc = str(row.get(c)).strip()
                break
        pay_entries.append({
            'date': row['date'].strftime('%d-%m-%Y') if pd.notna(row.get('date')) else '',
            'sort': row['date'] if pd.notna(row.get('date')) else None,
            'source': 'KVB', 'particulars': desc,
            'vendor': str(row.get('Client/Vendor', '') or ''),
            'amount': float(row.get('CR Amount', 0)),
        })
    for c in cash_payments:
        pay_entries.append({
            'date': (c.get('payment_date') or c.get('created_at') or '')[:10],
            'sort': c.get('payment_date') or c.get('created_at') or '',
            'source': 'Cash', 'particulars': c.get('note') or '',
            'vendor': '', 'amount': float(c.get('amount') or 0),
        })
    pay_entries.sort(key=lambda x: str(x['sort'] or ''), reverse=True)
    rr = hr + 1
    for e in pay_entries:
        ws.cell(row=rr, column=1, value=e['date'])
        ws.cell(row=rr, column=2, value=e['source']).font = (
            st['bank_kvb_font'] if e['source'] == 'KVB' else Font(bold=True, color='6B7280'))
        ws.cell(row=rr, column=3, value=e['particulars'])
        ws.cell(row=rr, column=4, value=e['vendor'])
        ac = ws.cell(row=rr, column=5, value=e['amount'])
        ac.number_format = currency_fmt
        ac.font = st['income_font']
        rr += 1
    ws.cell(row=rr, column=1, value='TOTAL').font = Font(bold=True)
    tc = ws.cell(row=rr, column=5, value=received_total)
    tc.font = Font(bold=True)
    tc.number_format = currency_fmt
    for col, w in (('A', 13), ('B', 10), ('C', 40), ('D', 24), ('E', 16)):
        ws.column_dimensions[col].width = w

    # ──────────────────────────────────────────────── SHEET 3: Expenses ──
    # Consolidated analytical sheet: total, a category breakdown, vendor + bank
    # breakdowns, and a filterable transactions table (Excel column auto-filter).
    ws = wb.create_sheet('Expenses')
    ws.cell(row=1, column=1, value=f'Expenses — {display}').font = st['title_font']
    if date_label:
        ws.cell(row=2, column=1, value=date_label).font = Font(italic=True, color='6B7280')
    expense_total = float(expense_df['DR Amount'].sum()) if not expense_df.empty else 0.0
    kv_row(ws, 3, 'Total Spent', expense_total, fmt=currency_fmt, font=st['expense_font'])
    ws.cell(row=3, column=3, value=f'{len(expense_df)} transaction(s)').font = Font(italic=True, color='6B7280')

    has_cats = (not expense_df.empty) and ('Category' in expense_df.columns)

    # ── BY CATEGORY ──
    r = 5
    ws.cell(row=r, column=1, value='BY CATEGORY').font = st['subtitle_font']
    r += 1
    for ci, h in enumerate(['Category', 'Amount'], 1):
        ws.cell(row=r, column=ci, value=h)
    style_header_row(ws, r, 2)
    r += 1
    if has_cats:
        cat_grp = expense_df.groupby('Category')['DR Amount'].sum().sort_values(ascending=False)
        for cat, amt in cat_grp.items():
            ws.cell(row=r, column=1, value=str(cat))
            ws.cell(row=r, column=2, value=float(amt)).number_format = currency_fmt
            r += 1
    ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
    tc = ws.cell(row=r, column=2, value=expense_total)
    tc.font = Font(bold=True)
    tc.number_format = currency_fmt
    r += 1

    # ── BY VENDOR ──
    r += 1
    ws.cell(row=r, column=1, value='BY VENDOR').font = st['subtitle_font']
    r += 1
    for ci, h in enumerate(['Vendor', 'Amount'], 1):
        ws.cell(row=r, column=ci, value=h)
    style_header_row(ws, r, 2)
    r += 1
    if not expense_df.empty and v_col in expense_df.columns:
        ven_grp = expense_df.groupby(v_col)['DR Amount'].sum().sort_values(ascending=False)
        for vn, amt in ven_grp.items():
            ws.cell(row=r, column=1, value=str(vn) if str(vn) != 'nan' else 'Unknown')
            ws.cell(row=r, column=2, value=float(amt)).number_format = currency_fmt
            r += 1

    # ── BY BANK ──
    r += 1
    ws.cell(row=r, column=1, value='BY BANK').font = st['subtitle_font']
    r += 1
    for ci, h in enumerate(['Bank', 'Amount'], 1):
        ws.cell(row=r, column=ci, value=h)
    style_header_row(ws, r, 2)
    r += 1
    if not expense_df.empty and 'bank' in expense_df.columns:
        bank_grp = expense_df.groupby('bank')['DR Amount'].sum().sort_values(ascending=False)
        for bcode, amt in bank_grp.items():
            bcell = ws.cell(row=r, column=1, value=str(bcode).upper())
            bf = bank_font(str(bcode))
            if bf:
                bcell.font = bf
            ws.cell(row=r, column=2, value=float(amt)).number_format = currency_fmt
            r += 1

    # ── TRANSACTIONS (filterable) ──
    r += 1
    ws.cell(row=r, column=1, value='TRANSACTIONS').font = st['subtitle_font']
    r += 1
    txn_header_row = r
    headers = ['Date', 'Paid To', 'Category', 'Bank', 'Amount']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    if not expense_df.empty:
        for _, row in expense_df.sort_values('date', ascending=False).iterrows():
            vend = str(row.get(v_col, '') or '')
            desc = ''
            for c in ('Description', 'Transaction Description', 'transaction_description'):
                if c in row and pd.notna(row.get(c)) and str(row.get(c)).strip():
                    desc = str(row.get(c)).strip()
                    break
            paid_to = vend if vend and vend.lower() not in ('unknown', 'nan', '') else desc
            bcode = str(row.get('bank', '') or '')
            ws.cell(row=r, column=1, value=row['date'].strftime('%d-%m-%Y') if pd.notna(row.get('date')) else '')
            ws.cell(row=r, column=2, value=paid_to)
            ws.cell(row=r, column=3, value=str(row.get('Category', '') or ''))
            bcell = ws.cell(row=r, column=4, value=bcode.upper())
            bf = bank_font(bcode)
            if bf:
                bcell.font = bf
            ac = ws.cell(row=r, column=5, value=float(row.get('DR Amount', 0)))
            ac.number_format = currency_fmt
            ac.font = st['expense_font']
            r += 1
    txn_last_row = r - 1
    # Excel column auto-filter over the transactions table.
    ws.auto_filter.ref = f'A{txn_header_row}:E{max(txn_last_row, txn_header_row)}'

    for col, w in (('A', 14), ('B', 34), ('C', 22), ('D', 10), ('E', 16)):
        ws.column_dimensions[col].width = w

    # ──────────────────────────────────────────── SHEET 4: Purchase Bills ──
    if purchase_bills:
        _write_project_bills_sheet(
            wb.create_sheet('Purchase Bills'), purchase_bills, stem_name,
            sheet_title=f'Purchase Bills — {display}',
            party_label='VENDOR', party_key='vendor_name', gstin_key='vendor_gstin',
            total_font=st['red_amount'], st=st)
    else:
        ws = wb.create_sheet('Purchase Bills')
        ws.cell(row=1, column=1, value='No purchase bills tagged to this project.')

    # ─────────────────────────────────────────────── SHEET 5: Sales Bills ──
    if sales_bills:
        _write_project_bills_sheet(
            wb.create_sheet('Sales Bills'), sales_bills, stem_name,
            sheet_title=f'Sales Bills — {display}',
            party_label='BUYER', party_key='buyer_name', gstin_key='buyer_gstin',
            total_font=st['green_amount'], st=st)
    else:
        ws = wb.create_sheet('Sales Bills')
        ws.cell(row=1, column=1, value='No sales bills tagged to this project.')

    # ───────────────────────────────────────────── SHEET 6: Labour Payments ──
    ws = wb.create_sheet('Labour Payments')
    ws.cell(row=1, column=1, value=f'Labour Payments — {display}').font = st['title_font']
    if not labour_summary.get('available', False):
        ws.cell(row=3, column=1, value=(
            'Could not reach the attendance app database — labour charges are '
            'unavailable.')).font = Font(italic=True, color='6B7280')
    else:
        # Monthly summary (mirrors the pop-up Labour tab)
        r = 3
        ws.cell(row=r, column=1, value='MONTHLY SUMMARY').font = st['subtitle_font']
        r += 1
        for ci, h in enumerate(['Month', 'Present Days', 'Workers', 'OT Hours', 'Cost'], 1):
            ws.cell(row=r, column=ci, value=h)
        style_header_row(ws, r, 5)
        r += 1
        for m in labour_summary.get('monthly', []):
            ws.cell(row=r, column=1, value=m.get('label', ''))
            ws.cell(row=r, column=2, value=m.get('days', 0))
            ws.cell(row=r, column=3, value=m.get('workers', 0))
            ws.cell(row=r, column=4, value=m.get('ot_hours', 0))
            ws.cell(row=r, column=5, value=float(m.get('cost', 0))).number_format = currency_fmt
            r += 1
        ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
        tc = ws.cell(row=r, column=5, value=labour_total)
        tc.font = st['green_amount']
        tc.number_format = currency_fmt
        r += 2
        # Per-worker detail per month
        labour_currency = currency_fmt
        for month_data in labour_monthly_detail:
            ws.cell(row=r, column=1,
                    value=f"{month_data['month_name'].upper()} — WORKER DETAIL").font = st['subtitle_font']
            r += 1
            wheaders = ['S.No', 'NAME', 'DESIGNATION', 'PRESENT DAYS', 'OT HOURS',
                        'BASE/DAY', 'BASE PAY', 'OT PAY', 'TOTAL SALARY']
            for ci, h in enumerate(wheaders, 1):
                ws.cell(row=r, column=ci, value=h)
            style_header_row(ws, r, len(wheaders))
            r += 1
            sno = 1
            for w in month_data['workers']:
                ws.cell(row=r, column=1, value=sno)
                ws.cell(row=r, column=2, value=w['name'])
                ws.cell(row=r, column=3, value=w['designation'])
                ws.cell(row=r, column=4, value=w['present_days'])
                ws.cell(row=r, column=5, value=w['ot_hours'])
                ws.cell(row=r, column=6, value=w['base_salary_per_day']).number_format = labour_currency
                ws.cell(row=r, column=7, value=w['base_pay']).number_format = labour_currency
                ws.cell(row=r, column=8, value=w['ot_pay']).number_format = labour_currency
                ws.cell(row=r, column=9, value=w['total_salary']).number_format = labour_currency
                sno += 1
                r += 1
            ws.cell(row=r, column=3, value='Month total').font = Font(bold=True)
            tc = ws.cell(row=r, column=9, value=float(month_data.get('total_salary', 0)))
            tc.font = st['green_amount']
            tc.number_format = labour_currency
            r += 2
    for col, w in (('A', 16), ('B', 24), ('C', 16), ('D', 13), ('E', 12),
                   ('F', 12), ('G', 13), ('H', 12), ('I', 14)):
        ws.column_dimensions[col].width = w

    # ──────────────────────────────────────── SHEET 7: Consolidated Summary ──
    ws = wb.create_sheet('Consolidated Summary')
    NUM = 3
    cr = 1
    for c in range(1, NUM + 1):
        ws.cell(row=cr, column=c).fill = st['project_name_fill']
    ws.cell(row=cr, column=1, value=f'PROJECT :  {display.upper()}').font = st['project_name_font']
    cr += 2

    def section(ws, row, text):
        ws.cell(row=row, column=1, value=text).font = st['section_bold']
        for c in range(1, NUM + 1):
            ws.cell(row=row, column=c).fill = st['green_fill']
        return row + 1

    cr = section(ws, cr, 'PROJECT VALUE')
    kv_row(ws, cr, 'Project Basic Value', value_basic, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'GST', value_gst, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'Total Value', value_total, fmt=currency_fmt, font=st['blue_amount']); cr += 1
    kv_row(ws, cr, 'Received — Bank (KVB)', received_bank, fmt=currency_fmt, font=st['income_font']); cr += 1
    kv_row(ws, cr, 'Received — Cash', received_cash, fmt=currency_fmt, font=st['income_font']); cr += 1
    kv_row(ws, cr, 'Total Received', received_total, fmt=currency_fmt, font=st['income_font']); cr += 1
    kv_row(ws, cr, ('Excess Received' if balance < -0.5 else 'Current Balance'), abs(balance),
           fmt=currency_fmt, font=(st['red_amount'] if balance > 0.5 else st['green_amount'])); cr += 1
    kv_row(ws, cr, 'PO Value (contract)', po_value, fmt=currency_fmt); cr += 2

    cr = section(ws, cr, 'GST POSITION')
    kv_row(ws, cr, 'Purchase — Basic', purchase_taxable, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'Purchase — GST', purchase_gst, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'Sales — Basic', sales_taxable, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'Sales — GST', sales_gst, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, ('GST Credit (carried forward)' if gst_extra < -0.5 else 'GST Extra (payable)'),
           abs(gst_extra), fmt=currency_fmt,
           font=(st['green_amount'] if gst_extra < -0.5 else st['red_amount'])); cr += 2

    cr = section(ws, cr, 'SPEND COMPOSITION')
    kv_row(ws, cr, 'Material Purchase (bills)', material_total, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'Other Expense', other_expense_total, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'Labour Payment', labour_total, fmt=currency_fmt, font=st['blue_amount']); cr += 1
    kv_row(ws, cr, 'GST Payable', gst_extra_cost, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'Overhead', overhead, fmt=currency_fmt); cr += 1
    kv_row(ws, cr, 'Total Cost', spend_total, fmt=currency_fmt, font=st['expense_font']); cr += 2

    cr = section(ws, cr, 'NET POSITION')
    kv_row(ws, cr, 'Balance (Total Value − Total Cost)', profit, fmt=currency_fmt,
           font=(st['green_amount'] if profit >= 0 else st['red_amount'])); cr += 1
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22

    # Expand every sheet's columns to fit their content so the workbook opens
    # fully readable — no manual column-widening needed. Runs last so it sizes
    # against the final content of every sheet (overrides the per-sheet widths).
    for sheet in wb.worksheets:
        auto_width(sheet, cap=50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_stem = re.sub(r'[^A-Za-z0-9_-]+', '_', stem_name).strip('_') or f'project_{project_id}'
    filename = f"Project_{project_id}_{safe_stem}_{now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

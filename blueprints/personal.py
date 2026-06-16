"""Personal expense tracker: pages + /api/personal/* endpoints."""

import io
from datetime import datetime

from flask import (
    Blueprint, render_template, request, jsonify, redirect, url_for, send_file,
)

from config import Config, now_ist
from database import build_project_filter_sql
from extensions import db_manager
from helpers.formatting import format_indian_number
from auth import login_required

bp = Blueprint('personal', __name__)


@bp.route('/personal-tracker')
@login_required
def personal_tracker():
    """Render personal transaction tracker page"""
    return render_template('personal_tracker.html')


@bp.route('/personal-tracker/add')
@login_required
def add_expense_page():
    """Render add expense page"""
    return render_template('expense_form.html', transaction=None)


@bp.route('/personal-tracker/edit/<int:transaction_id>')
@login_required
def edit_expense_page(transaction_id):
    """Render edit expense page"""
    if not Config.USE_DATABASE:
        return render_template('expense_form.html', transaction=None)

    try:
        with db_manager.get_connection() as conn:
            query = """
            SELECT id, transaction_date, vendor, description, project, amount,
                   COALESCE(transaction_type, 'expense') as transaction_type, bank
            FROM personal_transactions
            WHERE id = %s
            """
            cursor = conn.cursor(dictionary=True)
            cursor.execute(query, (transaction_id,))
            row = cursor.fetchone()
            cursor.close()

        if row:
            transaction = {
                'id': row['id'],
                'date': row['transaction_date'].strftime('%Y-%m-%d'),
                'vendor': row['vendor'],
                'description': row['description'] or '',
                'project': row['project'] or 'General',
                'amount': float(row['amount']),
                'transaction_type': row['transaction_type'] or 'expense',
                'bank': row.get('bank')
            }
            return render_template('expense_form.html', transaction=transaction)
        else:
            # Transaction not found, redirect to add page
            return redirect(url_for('personal.add_expense_page'))
    except Exception as e:

        return redirect(url_for('personal.add_expense_page'))


@bp.route('/api/personal/transactions', methods=['GET'])
@login_required
def get_personal_transactions():
    """Get all personal transactions"""
    if not Config.USE_DATABASE:
        return jsonify({'transactions': []})

    # Get filter parameters
    project = request.args.get('project', 'All')
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    search = request.args.get('search', '').lower()
    transaction_type = request.args.get('type', 'All')

    try:
        query = """
        SELECT id, transaction_date, vendor, description, project, amount,
               COALESCE(transaction_type, 'expense') as transaction_type, bank, created_at
        FROM personal_transactions
        WHERE 1=1
        """
        params = []

        if project and project != 'All':
            # fuzzy=True: tracker entries are typed free-hand ("jamuna lunch"),
            # so a canonical selection also stem-matches the project name part.
            # This is the ONLY project filter that stays fuzzy by design.
            proj_cond = build_project_filter_sql('project', project, params, fuzzy=True)
            if proj_cond:
                query += f" AND {proj_cond}"

        if transaction_type and transaction_type != 'All':
            query += " AND COALESCE(transaction_type, 'expense') = %s"
            params.append(transaction_type)

        if start_date:
            query += " AND transaction_date >= %s"
            params.append(start_date)

        if end_date:
            query += " AND transaction_date <= %s"
            params.append(end_date)

        if search:
            query += " AND (LOWER(vendor) LIKE %s OR LOWER(description) LIKE %s)"
            params.extend([f'%{search}%', f'%{search}%'])

        query += " ORDER BY transaction_date DESC, created_at DESC"

        with db_manager.get_connection() as conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(query, params)
            rows = cursor.fetchall()
            cursor.close()

        transactions = []
        for row in rows:
            trans_type = row.get('transaction_type', 'expense') or 'expense'
            transactions.append({
                'id': row['id'],
                'date': row['transaction_date'].strftime('%Y-%m-%d'),
                'date_formatted': row['transaction_date'].strftime('%d %b %Y'),
                'vendor': row['vendor'],
                'description': row['description'] or '',
                'project': row['project'],
                'amount': float(row['amount']),
                'amount_formatted': format_indian_number(row['amount']),
                'transaction_type': trans_type,
                'bank': row.get('bank')
            })
        return jsonify({'transactions': transactions})
    except Exception as e:

        return jsonify({'transactions': []})


@bp.route('/api/personal/transactions', methods=['POST'])
@login_required
def add_personal_transaction():
    """Add a new personal transaction"""
    if not Config.USE_DATABASE:
        return jsonify({'error': 'Database not available'}), 503

    data = request.json
    transaction_date = data.get('date')
    vendor = data.get('vendor', '').strip()
    description = data.get('description', '').strip()
    project = data.get('project', 'General').strip() or 'General'
    amount = data.get('amount')
    transaction_type = data.get('transaction_type', 'expense').strip().lower()
    bank = data.get('bank')

    # Validate transaction_type
    if transaction_type not in ['expense', 'income']:
        transaction_type = 'expense'

    # Validate bank
    if bank and bank not in ['axis', 'kvb']:
        bank = None

    if not transaction_date or not vendor or amount is None:
        return jsonify({'error': 'Missing required fields (date, vendor, amount)'}), 400

    try:
        amount = float(amount)
        if amount <= 0:
            return jsonify({'error': 'Amount must be greater than 0'}), 400
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid amount'}), 400

    try:
        with db_manager.get_connection() as conn:
            query = """
            INSERT INTO personal_transactions (transaction_date, vendor, description, project, amount, transaction_type, bank)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            cursor = conn.cursor()
            cursor.execute(query, (transaction_date, vendor, description, project, amount, transaction_type, bank))
            conn.commit()
            new_id = cursor.lastrowid
            cursor.close()

        return jsonify({
            'success': True,
            'message': 'Transaction added successfully',
            'id': new_id
        })
    except Exception as e:

        return jsonify({'error': str(e)}), 500


@bp.route('/api/personal/transactions/<int:transaction_id>', methods=['PUT'])
@login_required
def update_personal_transaction(transaction_id):
    """Update a personal transaction"""
    if not Config.USE_DATABASE:
        return jsonify({'error': 'Database not available'}), 503

    data = request.json
    transaction_date = data.get('date')
    vendor = data.get('vendor', '').strip()
    description = data.get('description', '').strip()
    project = data.get('project', 'General').strip() or 'General'
    amount = data.get('amount')
    transaction_type = data.get('transaction_type', 'expense').strip().lower()
    bank = data.get('bank')

    # Validate transaction_type
    if transaction_type not in ['expense', 'income']:
        transaction_type = 'expense'

    # Validate bank
    if bank and bank not in ['axis', 'kvb']:
        bank = None

    if not transaction_date or not vendor or amount is None:
        return jsonify({'error': 'Missing required fields (date, vendor, amount)'}), 400

    try:
        amount = float(amount)
        if amount <= 0:
            return jsonify({'error': 'Amount must be greater than 0'}), 400
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid amount'}), 400

    try:
        with db_manager.get_connection() as conn:
            query = """
            UPDATE personal_transactions
            SET transaction_date = %s, vendor = %s, description = %s, project = %s, amount = %s,
                transaction_type = %s, bank = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            """
            cursor = conn.cursor()
            cursor.execute(query, (transaction_date, vendor, description, project, amount, transaction_type, bank, transaction_id))
            conn.commit()
            affected_rows = cursor.rowcount
            cursor.close()

        if affected_rows > 0:
            return jsonify({
                'success': True,
                'message': 'Transaction updated successfully'
            })
        else:
            return jsonify({'error': 'Transaction not found'}), 404
    except Exception as e:

        return jsonify({'error': str(e)}), 500


@bp.route('/api/personal/transactions/<int:transaction_id>', methods=['DELETE'])
@login_required
def delete_personal_transaction(transaction_id):
    """Delete a personal transaction"""
    if not Config.USE_DATABASE:
        return jsonify({'error': 'Database not available'}), 503

    try:
        with db_manager.get_connection() as conn:
            query = "DELETE FROM personal_transactions WHERE id = %s"
            cursor = conn.cursor()
            cursor.execute(query, (transaction_id,))
            conn.commit()
            affected_rows = cursor.rowcount
            cursor.close()

        if affected_rows > 0:
            return jsonify({
                'success': True,
                'message': 'Transaction deleted successfully'
            })
        else:
            return jsonify({'error': 'Transaction not found'}), 404
    except Exception as e:

        return jsonify({'error': str(e)}), 500


@bp.route('/api/personal/export')
@login_required
def export_personal_transactions():
    """Export personal transactions as Excel"""
    if not Config.USE_DATABASE:
        return jsonify({'error': 'Database not available'}), 503

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        with db_manager.get_connection() as conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT transaction_date, vendor, description, project, amount,
                       COALESCE(transaction_type, 'expense') as transaction_type, bank
                FROM personal_transactions
                ORDER BY transaction_date DESC, created_at DESC
            """)
            rows = cursor.fetchall()
            cursor.close()

        wb = Workbook()

        # ── Styles ──
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        currency_fmt = '#,##0.00'
        date_fmt = 'DD-MMM-YYYY'
        income_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
        expense_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

        # ── All Transactions sheet ──
        ws = wb.active
        ws.title = 'All Transactions'
        headers = ['Date', 'Type', 'Vendor', 'Description', 'Project', 'Bank', 'Amount']
        col_widths = [14, 10, 25, 30, 20, 10, 15]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, row in enumerate(rows, 2):
            trans_type = row.get('transaction_type', 'expense') or 'expense'
            row_fill = income_fill if trans_type == 'income' else expense_fill

            values = [
                row['transaction_date'],
                trans_type.capitalize(),
                row['vendor'],
                row['description'] or '',
                row['project'],
                (row.get('bank') or '').upper(),
                float(row['amount'])
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.fill = row_fill
                if col_idx == 1:
                    cell.number_format = date_fmt
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx == 7:
                    cell.number_format = currency_fmt
                    cell.alignment = Alignment(horizontal='right')

        ws.auto_filter.ref = f"A1:G{max(len(rows) + 1, 2)}"
        ws.freeze_panes = 'A2'

        # ── Monthly Summary sheet ──
        ws_monthly = wb.create_sheet('Monthly Summary')
        monthly = {}
        for row in rows:
            key = row['transaction_date'].strftime('%Y-%m')
            if key not in monthly:
                monthly[key] = {'income': 0, 'expense': 0}
            trans_type = row.get('transaction_type', 'expense') or 'expense'
            monthly[key][trans_type] += float(row['amount'])

        m_headers = ['Month', 'Income', 'Expenses', 'Net']
        m_widths = [16, 15, 15, 15]
        for col_idx, (header, width) in enumerate(zip(m_headers, m_widths), 1):
            cell = ws_monthly.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws_monthly.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, (month_key, totals) in enumerate(sorted(monthly.items(), reverse=True), 2):
            month_label = datetime.strptime(month_key, '%Y-%m').strftime('%B %Y')
            net = totals['income'] - totals['expense']
            values = [month_label, totals['income'], totals['expense'], net]
            for col_idx, val in enumerate(values, 1):
                cell = ws_monthly.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx >= 2:
                    cell.number_format = currency_fmt
                    cell.alignment = Alignment(horizontal='right')

        # ── Project Summary sheet ──
        ws_proj = wb.create_sheet('Project Summary')
        proj_totals = {}
        for row in rows:
            proj = row['project']
            trans_type = row.get('transaction_type', 'expense') or 'expense'
            if proj not in proj_totals:
                proj_totals[proj] = {'income': 0, 'expense': 0, 'count': 0}
            proj_totals[proj][trans_type] += float(row['amount'])
            proj_totals[proj]['count'] += 1

        p_headers = ['Project', 'Income', 'Expenses', 'Net', 'Transactions']
        p_widths = [25, 15, 15, 15, 14]
        for col_idx, (header, width) in enumerate(zip(p_headers, p_widths), 1):
            cell = ws_proj.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws_proj.column_dimensions[get_column_letter(col_idx)].width = width

        sorted_projs = sorted(proj_totals.items(), key=lambda x: x[1]['expense'], reverse=True)
        for row_idx, (proj, totals) in enumerate(sorted_projs, 2):
            net = totals['income'] - totals['expense']
            values = [proj, totals['income'], totals['expense'], net, totals['count']]
            for col_idx, val in enumerate(values, 1):
                cell = ws_proj.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (2, 3, 4):
                    cell.number_format = currency_fmt
                    cell.alignment = Alignment(horizontal='right')
                elif col_idx == 5:
                    cell.alignment = Alignment(horizontal='center')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Expense_Tracker_{now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/api/personal/summary')
@login_required
def get_personal_summary():
    """Get summary statistics for personal transactions"""
    empty_response = {
        'total_expense': 0,
        'total_expense_formatted': '₹0',
        'total_income': 0,
        'total_income_formatted': '₹0',
        'net_balance': 0,
        'net_balance_formatted': '₹0',
        'this_month_expense': 0,
        'this_month_expense_formatted': '₹0',
        'this_month_income': 0,
        'this_month_income_formatted': '₹0',
        'transaction_count': 0,
        'project_breakdown': []
    }

    if not Config.USE_DATABASE:
        return jsonify(empty_response)

    # Get filter parameters
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)

    try:
        with db_manager.get_connection() as conn:
            cursor = conn.cursor(dictionary=True)

            # Total expense and income query
            total_query = """
            SELECT
                COALESCE(SUM(CASE WHEN COALESCE(transaction_type, 'expense') = 'expense' THEN amount ELSE 0 END), 0) as total_expense,
                COALESCE(SUM(CASE WHEN COALESCE(transaction_type, 'expense') = 'income' THEN amount ELSE 0 END), 0) as total_income,
                COUNT(*) as count
            FROM personal_transactions
            WHERE 1=1
            """
            params = []

            if start_date:
                total_query += " AND transaction_date >= %s"
                params.append(start_date)
            if end_date:
                total_query += " AND transaction_date <= %s"
                params.append(end_date)

            cursor.execute(total_query, params)
            total_result = cursor.fetchone()
            total_expense = float(total_result['total_expense']) if total_result else 0
            total_income = float(total_result['total_income']) if total_result else 0
            net_balance = total_income - total_expense
            transaction_count = int(total_result['count']) if total_result else 0

            # This month query with income/expense
            this_month_query = """
            SELECT
                COALESCE(SUM(CASE WHEN COALESCE(transaction_type, 'expense') = 'expense' THEN amount ELSE 0 END), 0) as expense,
                COALESCE(SUM(CASE WHEN COALESCE(transaction_type, 'expense') = 'income' THEN amount ELSE 0 END), 0) as income
            FROM personal_transactions
            WHERE YEAR(transaction_date) = YEAR(CURRENT_DATE)
              AND MONTH(transaction_date) = MONTH(CURRENT_DATE)
            """
            cursor.execute(this_month_query)
            this_month_result = cursor.fetchone()
            this_month_expense = float(this_month_result['expense']) if this_month_result else 0
            this_month_income = float(this_month_result['income']) if this_month_result else 0

            # Project breakdown query (expenses only)
            project_query = """
            SELECT project, SUM(amount) as total, COUNT(*) as count
            FROM personal_transactions
            WHERE COALESCE(transaction_type, 'expense') = 'expense'
            """
            params = []
            if start_date:
                project_query += " AND transaction_date >= %s"
                params.append(start_date)
            if end_date:
                project_query += " AND transaction_date <= %s"
                params.append(end_date)

            project_query += " GROUP BY project ORDER BY total DESC"

            cursor.execute(project_query, params)
            project_rows = cursor.fetchall()
            cursor.close()

        project_breakdown = []
        for row in project_rows:
            pct = (float(row['total']) / total_expense * 100) if total_expense > 0 else 0
            project_breakdown.append({
                'project': row['project'],
                'amount': float(row['total']),
                'amount_formatted': format_indian_number(row['total']),
                'count': int(row['count']),
                'percentage': round(pct, 1)
            })

        return jsonify({
            'total_expense': total_expense,
            'total_expense_formatted': format_indian_number(total_expense),
            'total_income': total_income,
            'total_income_formatted': format_indian_number(total_income),
            'net_balance': net_balance,
            'net_balance_formatted': format_indian_number(abs(net_balance)),
            'net_balance_positive': net_balance >= 0,
            'this_month_expense': this_month_expense,
            'this_month_expense_formatted': format_indian_number(this_month_expense),
            'this_month_income': this_month_income,
            'this_month_income_formatted': format_indian_number(this_month_income),
            'transaction_count': transaction_count,
            'project_breakdown': project_breakdown
        })
    except Exception as e:

        return jsonify(empty_response)


@bp.route('/api/personal/projects')
@login_required
def get_personal_projects():
    """Get list of unique projects from personal transactions"""
    if not Config.USE_DATABASE:
        return jsonify({'projects': ['General']})

    try:
        with db_manager.get_connection() as conn:
            query = "SELECT DISTINCT project FROM personal_transactions ORDER BY project"
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            cursor.close()

        projects = [row[0] for row in rows if row[0]]
        if not projects:
            projects = ['General']
        return jsonify({'projects': projects})
    except Exception as e:

        return jsonify({'projects': ['General']})


@bp.route('/api/personal/vendors')
@login_required
def get_personal_vendors():
    """Get list of unique vendors from personal transactions"""
    if not Config.USE_DATABASE:
        return jsonify({'vendors': []})

    try:
        with db_manager.get_connection() as conn:
            query = "SELECT DISTINCT vendor FROM personal_transactions ORDER BY vendor"
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            cursor.close()

        vendors = [row[0] for row in rows if row[0]]
        return jsonify({'vendors': vendors})
    except Exception as e:

        return jsonify({'vendors': []})


@bp.route('/api/personal/descriptions')
@login_required
def get_personal_descriptions():
    """Get list of unique descriptions from personal transactions"""
    if not Config.USE_DATABASE:
        return jsonify({'descriptions': []})

    try:
        with db_manager.get_connection() as conn:
            query = "SELECT DISTINCT description FROM personal_transactions WHERE description IS NOT NULL AND description != '' ORDER BY description"
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            cursor.close()

        descriptions = [row[0] for row in rows if row[0]]
        return jsonify({'descriptions': descriptions})
    except Exception as e:

        return jsonify({'descriptions': []})



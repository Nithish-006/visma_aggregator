from flask import Flask, render_template, jsonify, request, send_file, redirect, url_for, session
from functools import wraps
import pandas as pd
import json
import io
import os
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import re

# Import our modules
from config import Config, allowed_file, BANK_CONFIG, VALID_BANK_CODES, get_bank_config, get_bank_table, now_ist
from database import DatabaseManager, build_project_filter_sql
from bank_statement_processor import process_bank_statement
from bill_processor import process_bill_file, generate_excel, format_extracted_data_for_display
from extraction_validator import validate_extraction, validate_db_row
import po_processor

# Shared singletons / mutable state (extracted from this file's old globals)
from extensions import db_manager, state
from helpers.formatting import sanitize_for_excel, safe_col_width, format_indian_number
from helpers.projects import (
    get_project_stems, normalize_project_stem, build_smart_project_groups,
    match_bills_to_project_groups, match_labour_to_project_groups,
    parse_project_selection, project_value_matches_selection, validate_project_value,
)
from helpers.bankdata import load_bank_data_from_db, get_bank_df, reload_bank_data
from helpers.dataframe import (
    load_financial_data_from_db, load_financial_data_from_excel, reload_data,
    parse_month_filter, filter_by_months, filter_by_date_range,
    filter_by_project, filter_by_category, filter_by_vendor, robust_filter_by_project,
)

app = Flask(__name__)
app.config.from_object(Config)

# Secret key for session management
app.secret_key = os.environ.get('SECRET_KEY', 'visma-finance-secret-key-2024-secure')

# Permanent session lifetime (30 days for "Stay signed in")
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

# Auth lives in its own blueprint; login_required is imported back for the
# routes still defined in this module.
from auth import bp as auth_bp, login_required
from blueprints.personal import bp as personal_bp
from blueprints.sales import bp as sales_bp
from blueprints.bills import bp as bills_bp
from blueprints.banks import bp as banks_bp
from blueprints.legacy import bp as legacy_bp
from blueprints.projects import bp as projects_bp
from blueprints.project_summary import bp as project_summary_bp
app.register_blueprint(auth_bp)
app.register_blueprint(personal_bp)
app.register_blueprint(sales_bp)
app.register_blueprint(bills_bp)
app.register_blueprint(banks_bp)
app.register_blueprint(legacy_bp)
app.register_blueprint(projects_bp)
app.register_blueprint(project_summary_bp)

# Create uploads directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


# Load legacy data at startup (populates extensions.state.df_global)
reload_data()


# ============================================================================
# PROJECT SUMMARY - Cross-Bank Consolidated View
# ============================================================================


@app.route('/api/project-summary/export')
@login_required
def export_project_summary():
    """Export professional project summary report as multi-tab Excel"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    project = request.args.get('project', None)
    category = request.args.get('category', None)
    vendor = request.args.get('vendor', None)

    # Gather combined data
    combined_rows = []
    for bank_code in VALID_BANK_CODES:
        df = get_bank_df(bank_code).copy()
        if df.empty:
            continue
        df = filter_by_date_range(df, start_date, end_date)
        df = robust_filter_by_project(df, project)
        df = filter_by_category(df, category)
        df = filter_by_vendor(df, vendor)
        if df.empty:
            continue
        df['bank'] = bank_code
        combined_rows.append(df)

    if not combined_rows:
        combined = pd.DataFrame()
    else:
        combined = pd.concat(combined_rows, ignore_index=True)

    # Style constants
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='1A1A2E')
    subtitle_font = Font(name='Calibri', bold=True, size=11, color='4A4A68')
    currency_fmt = '#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB')
    )
    income_font = Font(name='Calibri', color='059669', bold=True)
    expense_font = Font(name='Calibri', color='DC2626', bold=True)
    axis_fill = PatternFill(start_color='FDF2F8', end_color='FDF2F8', fill_type='solid')
    kvb_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    val = str(cell.value) if cell.value else ''
                    max_len = max(max_len, len(val))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ──────────────────────────────────────────────────────────
        # TAB 1: Executive Summary
        # ──────────────────────────────────────────────────────────
        kvb_export_df = combined[combined['bank'] == 'kvb'] if not combined.empty else pd.DataFrame()
        axis_export_df = combined[combined['bank'] == 'axis'] if not combined.empty else pd.DataFrame()
        total_income = float(kvb_export_df['CR Amount'].sum()) if not kvb_export_df.empty else 0
        total_bank_transfer = float(axis_export_df['CR Amount'].sum()) if not axis_export_df.empty else 0
        total_credit_all = float(combined['CR Amount'].sum()) if not combined.empty else 0
        total_expense = float(combined['DR Amount'].sum()) if not combined.empty else 0
        txn_count = len(combined) if not combined.empty else 0

        date_label = ''
        if start_date and end_date:
            date_label = f"{start_date} to {end_date}"
        elif start_date:
            date_label = f"From {start_date}"
        elif end_date:
            date_label = f"Up to {end_date}"
        else:
            date_label = 'All Time'

        summary_data = [
            ['VISMA Financial - Project Summary Report'],
            [''],
            ['Report Period', date_label],
            ['Generated On', now_ist().strftime('%d-%b-%Y %H:%M')],
            [''],
            ['KEY PERFORMANCE INDICATORS'],
            [''],
            ['Metric', 'Value'],
            ['Total Income (KVB)', total_income],
            ['Total Bank Transfer (Axis)', total_bank_transfer],
            ['Total Expense', total_expense],
            ['Total Transactions', txn_count],
            ['Expense Ratio', total_expense / total_income if total_income > 0 else 0],
            ['Average Transaction Size', (total_income + total_bank_transfer + total_expense) / txn_count if txn_count > 0 else 0],
        ]

        # Per-bank KPIs
        summary_data.append([''])
        summary_data.append(['BANK-WISE SUMMARY'])
        summary_data.append([''])
        summary_data.append(['Bank', 'Income / Bank Transfer', 'Expense', 'Net', 'Transactions', '% of Total Expense'])

        for bc in VALID_BANK_CODES:
            if combined.empty:
                continue
            bdf = combined[combined['bank'] == bc]
            if bdf.empty:
                continue
            b_inc = float(bdf['CR Amount'].sum())
            b_exp = float(bdf['DR Amount'].sum())
            b_net = b_inc - b_exp
            b_cnt = len(bdf)
            b_pct = b_exp / total_expense if total_expense > 0 else 0
            bank_name = get_bank_config(bc)['name']
            summary_data.append([bank_name, b_inc, b_exp, b_net, b_cnt, b_pct])

        # Active filters
        summary_data.append([''])
        summary_data.append(['APPLIED FILTERS'])
        summary_data.append([''])
        if project:
            summary_data.append(['Project Filter', project])
        if category:
            summary_data.append(['Category Filter', category])
        if vendor:
            summary_data.append(['Vendor Filter', vendor])
        if not project and not category and not vendor:
            summary_data.append(['Filters', 'None (all data)'])

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Executive Summary', index=False, header=False)

        ws = writer.sheets['Executive Summary']
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=6, column=1).font = subtitle_font
        ws.cell(row=8, column=1).font = header_font
        ws.cell(row=8, column=1).fill = header_fill
        ws.cell(row=8, column=2).font = header_font
        ws.cell(row=8, column=2).fill = header_fill

        # Format currency cells
        for r in range(9, 15):
            cell = ws.cell(row=r, column=2)
            if r == 12:  # Transaction count
                pass
            elif r == 13:  # Expense ratio
                cell.number_format = pct_fmt
            else:
                cell.number_format = currency_fmt
            if r == 9:   # Total Income (KVB)
                cell.font = income_font
            elif r == 10:  # Total Bank Transfer (Axis)
                cell.font = Font(color='2563EB', bold=True)
            elif r == 11:  # Total Expense
                cell.font = expense_font

        # Style bank summary header
        bank_header_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
            for cell in row:
                if cell.value == 'Bank':
                    bank_header_row = cell.row
                    break
        if bank_header_row:
            style_header_row(ws, bank_header_row, 6)
            for r in range(bank_header_row + 1, ws.max_row + 1):
                c = ws.cell(row=r, column=1)
                if c.value and ('Axis' in str(c.value) or 'KVB' in str(c.value) or 'Karur' in str(c.value)):
                    for col in range(2, 5):
                        ws.cell(row=r, column=col).number_format = currency_fmt
                    ws.cell(row=r, column=6).number_format = pct_fmt

        # Style filters header
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
            for cell in row:
                if cell.value == 'BANK-WISE SUMMARY' or cell.value == 'APPLIED FILTERS':
                    cell.font = subtitle_font

        auto_width(ws)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # ──────────────────────────────────────────────────────────
        # TAB 2: Expense Breakdown (by Category)
        # ──────────────────────────────────────────────────────────
        if not combined.empty:
            expense_df = combined[combined['DR Amount'] > 0]
            if not expense_df.empty:
                cat_data = expense_df.groupby('Category').agg(
                    Total_Expense=('DR Amount', 'sum')
                ).sort_values('Total_Expense', ascending=False).reset_index()

                total_exp = cat_data['Total_Expense'].sum()

                cat_data.columns = ['Category', 'Total Expense']

                cat_data.to_excel(writer, sheet_name='Expense Breakdown', index=False, startrow=2)
                ws2 = writer.sheets['Expense Breakdown']
                ws2.cell(row=1, column=1, value='Expense Breakdown by Category').font = title_font
                style_header_row(ws2, 3, 2)

                for r in range(4, 4 + len(cat_data)):
                    ws2.cell(row=r, column=2).number_format = currency_fmt

                # Add total row
                total_row = 4 + len(cat_data)
                ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
                ws2.cell(row=total_row, column=2, value=total_exp).font = Font(bold=True)
                ws2.cell(row=total_row, column=2).number_format = currency_fmt

                auto_width(ws2)
            else:
                pd.DataFrame({'Note': ['No expense data for the selected filters']}).to_excel(
                    writer, sheet_name='Expense Breakdown', index=False)

        # ──────────────────────────────────────────────────────────
        # TAB 3: Cashflow Analysis (by Project)
        # ──────────────────────────────────────────────────────────
        if not combined.empty:
            project_col = 'Project' if 'Project' in combined.columns else 'project'
            if project_col in combined.columns:
                proj_income = combined.groupby(project_col)['CR Amount'].sum()
                proj_expense = combined.groupby(project_col)['DR Amount'].sum()
                all_projects_list = sorted(set(proj_income.index) | set(proj_expense.index))

                proj_rows = []
                for p in all_projects_list:
                    p_inc = float(proj_income.get(p, 0))
                    p_exp = float(proj_expense.get(p, 0))
                    proj_rows.append({
                        'Project': str(p) if p and str(p) != 'nan' else 'Unassigned',
                        'Income': p_inc, 'Expense': p_exp
                    })

                df_proj = pd.DataFrame(proj_rows).sort_values('Expense', ascending=False)
                df_proj.to_excel(writer, sheet_name='Cashflow Analysis', index=False, startrow=2)

                ws3 = writer.sheets['Cashflow Analysis']
                ws3.cell(row=1, column=1, value='Project Cashflow Analysis').font = title_font
                style_header_row(ws3, 3, 3)

                for r in range(4, 4 + len(df_proj)):
                    for c in [2, 3]:
                        ws3.cell(row=r, column=c).number_format = currency_fmt

                # Add total row
                total_row = 4 + len(df_proj)
                ws3.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
                ws3.cell(row=total_row, column=2, value=total_credit_all).font = Font(bold=True)
                ws3.cell(row=total_row, column=2).number_format = currency_fmt
                ws3.cell(row=total_row, column=3, value=total_expense).font = Font(bold=True)
                ws3.cell(row=total_row, column=3).number_format = currency_fmt

                auto_width(ws3)

        # ──────────────────────────────────────────────────────────
        # TAB 4: Vendor Breakdown
        # ──────────────────────────────────────────────────────────
        if not combined.empty:
            expense_df = combined[combined['DR Amount'] > 0]
            vendor_col = 'Client/Vendor' if 'Client/Vendor' in combined.columns else 'client_vendor'
            if vendor_col in combined.columns and not expense_df.empty:
                vendor_data = expense_df.groupby(vendor_col).agg(
                    Total_Expense=('DR Amount', 'sum')
                ).sort_values('Total_Expense', ascending=False).reset_index()

                total_v_exp = vendor_data['Total_Expense'].sum()

                vendor_data.columns = ['Vendor', 'Total Expense']

                vendor_data.to_excel(writer, sheet_name='Vendor Breakdown', index=False, startrow=2)

                ws4 = writer.sheets['Vendor Breakdown']
                ws4.cell(row=1, column=1, value='Vendor Expense Breakdown').font = title_font
                style_header_row(ws4, 3, 2)

                for r in range(4, 4 + len(vendor_data)):
                    ws4.cell(row=r, column=2).number_format = currency_fmt

                # Total row
                total_row = 4 + len(vendor_data)
                ws4.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
                ws4.cell(row=total_row, column=2, value=total_v_exp).font = Font(bold=True)
                ws4.cell(row=total_row, column=2).number_format = currency_fmt

                auto_width(ws4)
            else:
                pd.DataFrame({'Note': ['No vendor expense data for the selected filters']}).to_excel(
                    writer, sheet_name='Vendor Breakdown', index=False)

        # ──────────────────────────────────────────────────────────
        # TAB 5 & 6: Bank-wise Transactions (one tab per bank)
        # ──────────────────────────────────────────────────────────
        for bc in VALID_BANK_CODES:
            bank_config = get_bank_config(bc)
            sheet_name = f"{bank_config['name']} Txns"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            if combined.empty:
                pd.DataFrame({'Note': ['No data']}).to_excel(writer, sheet_name=sheet_name, index=False)
                continue

            bdf = combined[combined['bank'] == bc].sort_values('date', ascending=False).copy()
            if bdf.empty:
                pd.DataFrame({'Note': [f'No transactions for {bank_config["name"]}']}).to_excel(
                    writer, sheet_name=sheet_name, index=False)
                continue

            bdf['Date'] = bdf['date'].dt.strftime('%d-%m-%Y')
            export_cols = {
                'Date': 'Date',
                'Client/Vendor': 'Vendor',
                'Category': 'Category',
                'Project': 'Project',
                'DR Amount': 'Debit (₹)',
                'CR Amount': 'Credit (₹)'
            }

            df_export = pd.DataFrame()
            for src, dst in export_cols.items():
                if src in bdf.columns:
                    df_export[dst] = bdf[src]
                elif src.lower() in bdf.columns:
                    df_export[dst] = bdf[src.lower()]
                else:
                    df_export[dst] = None

            df_export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

            ws_b = writer.sheets[sheet_name]
            ws_b.cell(row=1, column=1, value=f'{bank_config["name"]} - Transaction Details').font = title_font
            style_header_row(ws_b, 3, len(df_export.columns))

            for r in range(4, 4 + len(df_export)):
                ws_b.cell(row=r, column=5).number_format = currency_fmt
                ws_b.cell(row=r, column=6).number_format = currency_fmt
                dr_val = ws_b.cell(row=r, column=5).value
                cr_val = ws_b.cell(row=r, column=6).value
                if dr_val and float(dr_val) > 0:
                    ws_b.cell(row=r, column=5).font = expense_font
                if cr_val and float(cr_val) > 0:
                    ws_b.cell(row=r, column=6).font = income_font

            # Bank subtotals
            total_row = 4 + len(df_export)
            ws_b.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
            ws_b.cell(row=total_row, column=5, value=float(bdf['DR Amount'].sum())).font = Font(bold=True)
            ws_b.cell(row=total_row, column=5).number_format = currency_fmt
            ws_b.cell(row=total_row, column=6, value=float(bdf['CR Amount'].sum())).font = Font(bold=True)
            ws_b.cell(row=total_row, column=6).number_format = currency_fmt

            auto_width(ws_b)

        # ──────────────────────────────────────────────────────────
        # TAB 7 & 8: Purchase Bills / Sales Bills (project-grouped)
        # ──────────────────────────────────────────────────────────
        # Shared auditor-format styles (reused by Project Breakdown tab too)
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        block_bg = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')  # mild yellow
        project_name_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        project_name_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
        separator_fill = PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid')

        # Shared helper: writes a project-grouped bills sheet
        def write_bills_sheet(ws, bills_by_stem_local, stem_groups_local,
                              sheet_title, party_label, party_key, gstin_key,
                              total_font):
            """Write a project-grouped bills sheet with expanded line items.

            party_label: 'VENDOR' or 'BUYER'
            party_key: 'vendor_name' or 'buyer_name'
            gstin_key: 'vendor_gstin' or 'buyer_gstin'
            total_font: font colour for totals (red_amount or green_amount-like)
            """
            BILL_COLS = 15  # SL.NO, Party, GSTIN, Invoice#, Date, Description, HSN/SAC, QTY, UOM, RATE, TAXABLE, CGST, SGST, IGST, TOTAL
            bill_header_labels = [
                'SL.NO', party_label, 'GSTIN', 'INVOICE #', 'DATE',
                'DESCRIPTION', 'HSN/SAC', 'QTY', 'UOM', 'RATE',
                'TAXABLE AMT', 'CGST', 'SGST', 'IGST', 'TOTAL'
            ]

            ws.cell(row=1, column=1, value=sheet_title).font = title_font
            cr = 3  # current row
            grand_taxable = 0
            grand_cgst = 0
            grand_sgst = 0
            grand_igst = 0
            grand_total = 0
            bill_serial = 0

            for stem in sorted(stem_groups_local.keys()):
                group_bills = bills_by_stem_local.get(stem, [])
                if not group_bills:
                    continue

                project_names = stem_groups_local[stem]
                group_label = stem.upper()
                proj_list = ', '.join(sorted(str(p) for p in project_names if str(p) != 'nan'))

                block_start = cr

                # ── PROJECT HEADER (blue fill, white text) ──
                for c in range(1, BILL_COLS + 1):
                    ws.cell(row=cr, column=c).fill = project_name_fill
                ws.cell(row=cr, column=1,
                        value=f'PROJECT :  {group_label}').font = project_name_font
                cr += 1

                # Variant names
                ws.cell(row=cr, column=1,
                        value=f'({proj_list})').font = Font(
                    name='Calibri', italic=True, color='6B7280', size=9)
                cr += 2

                # ── COLUMN HEADERS ──
                for ci, lbl in enumerate(bill_header_labels, 1):
                    cell = ws.cell(row=cr, column=ci, value=lbl)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cr += 1

                proj_taxable = 0
                proj_cgst = 0
                proj_sgst = 0
                proj_igst = 0
                proj_total = 0

                for bill in group_bills:
                    bill_serial += 1
                    line_items = bill.get('line_items', [])
                    b_taxable = float(bill.get('subtotal', 0) or 0)
                    b_cgst = float(bill.get('total_cgst', 0) or 0)
                    b_sgst = float(bill.get('total_sgst', 0) or 0)
                    b_igst = float(bill.get('total_igst', 0) or 0)
                    b_total = float(bill.get('total_amount', 0) or 0)

                    if line_items:
                        # First line item shares the row with bill header info
                        for li_idx, item in enumerate(line_items):
                            if li_idx == 0:
                                ws.cell(row=cr, column=1, value=bill_serial)
                                ws.cell(row=cr, column=2, value=bill.get(party_key, ''))
                                ws.cell(row=cr, column=3, value=bill.get(gstin_key, ''))
                                ws.cell(row=cr, column=4, value=bill.get('invoice_number', ''))
                                ws.cell(row=cr, column=5, value=bill.get('invoice_date', ''))
                            # Line item detail columns
                            ws.cell(row=cr, column=6, value=item.get('description', ''))
                            ws.cell(row=cr, column=7, value=item.get('hsn_sac_code', ''))
                            qty = item.get('quantity', 0)
                            if qty:
                                ws.cell(row=cr, column=8, value=qty)
                                ws.cell(row=cr, column=8).number_format = '#,##0.00'
                            ws.cell(row=cr, column=9, value=item.get('uom', ''))
                            rate = item.get('rate_per_unit', 0)
                            if rate:
                                ws.cell(row=cr, column=10, value=rate)
                                ws.cell(row=cr, column=10).number_format = currency_fmt
                            taxable = item.get('taxable_value', 0)
                            if taxable:
                                ws.cell(row=cr, column=11, value=taxable)
                                ws.cell(row=cr, column=11).number_format = currency_fmt
                            item_cgst = item.get('cgst_amount', 0)
                            if item_cgst:
                                ws.cell(row=cr, column=12, value=item_cgst)
                                ws.cell(row=cr, column=12).number_format = currency_fmt
                            item_sgst = item.get('sgst_amount', 0)
                            if item_sgst:
                                ws.cell(row=cr, column=13, value=item_sgst)
                                ws.cell(row=cr, column=13).number_format = currency_fmt
                            item_igst = item.get('igst_amount', 0)
                            if item_igst:
                                ws.cell(row=cr, column=14, value=item_igst)
                                ws.cell(row=cr, column=14).number_format = currency_fmt
                            cr += 1
                    else:
                        # Bill with no line items - single row with bill totals
                        ws.cell(row=cr, column=1, value=bill_serial)
                        ws.cell(row=cr, column=2, value=bill.get(party_key, ''))
                        ws.cell(row=cr, column=3, value=bill.get(gstin_key, ''))
                        ws.cell(row=cr, column=4, value=bill.get('invoice_number', ''))
                        ws.cell(row=cr, column=5, value=bill.get('invoice_date', ''))
                        ws.cell(row=cr, column=11, value=b_taxable)
                        ws.cell(row=cr, column=11).number_format = currency_fmt
                        ws.cell(row=cr, column=12, value=b_cgst)
                        ws.cell(row=cr, column=12).number_format = currency_fmt
                        ws.cell(row=cr, column=13, value=b_sgst)
                        ws.cell(row=cr, column=13).number_format = currency_fmt
                        ws.cell(row=cr, column=14, value=b_igst)
                        ws.cell(row=cr, column=14).number_format = currency_fmt
                        ws.cell(row=cr, column=15, value=b_total)
                        ws.cell(row=cr, column=15).number_format = currency_fmt
                        cr += 1

                    # ── Bill Total row ──
                    ws.cell(row=cr, column=6, value='Bill Total').font = Font(bold=True)
                    ws.cell(row=cr, column=11, value=b_taxable).font = Font(bold=True)
                    ws.cell(row=cr, column=11).number_format = currency_fmt
                    ws.cell(row=cr, column=12, value=b_cgst).font = Font(bold=True)
                    ws.cell(row=cr, column=12).number_format = currency_fmt
                    ws.cell(row=cr, column=13, value=b_sgst).font = Font(bold=True)
                    ws.cell(row=cr, column=13).number_format = currency_fmt
                    ws.cell(row=cr, column=14, value=b_igst).font = Font(bold=True)
                    ws.cell(row=cr, column=14).number_format = currency_fmt
                    ws.cell(row=cr, column=15, value=b_total).font = Font(bold=True)
                    ws.cell(row=cr, column=15).number_format = currency_fmt
                    # Thin bottom border on bill total row
                    for c in range(1, BILL_COLS + 1):
                        ws.cell(row=cr, column=c).border = thin_border
                    cr += 1

                    proj_taxable += b_taxable
                    proj_cgst += b_cgst
                    proj_sgst += b_sgst
                    proj_igst += b_igst
                    proj_total += b_total

                # ── PROJECT TOTAL (green fill) ──
                for c in range(1, BILL_COLS + 1):
                    ws.cell(row=cr, column=c).fill = green_fill
                ws.cell(row=cr, column=1, value=f'PROJECT TOTAL — {group_label}').font = Font(bold=True)
                ws.cell(row=cr, column=11, value=proj_taxable).font = total_font
                ws.cell(row=cr, column=11).number_format = currency_fmt
                ws.cell(row=cr, column=12, value=proj_cgst).font = total_font
                ws.cell(row=cr, column=12).number_format = currency_fmt
                ws.cell(row=cr, column=13, value=proj_sgst).font = total_font
                ws.cell(row=cr, column=13).number_format = currency_fmt
                ws.cell(row=cr, column=14, value=proj_igst).font = total_font
                ws.cell(row=cr, column=14).number_format = currency_fmt
                ws.cell(row=cr, column=15, value=proj_total).font = total_font
                ws.cell(row=cr, column=15).number_format = currency_fmt
                cr += 1

                # Yellow background on all content rows in this block
                for r in range(block_start, cr):
                    for c in range(1, BILL_COLS + 1):
                        cell = ws.cell(row=r, column=c)
                        if cell.fill == PatternFill(fill_type=None) or cell.fill == PatternFill():
                            cell.fill = block_bg

                # ── Dark separator ──
                for c in range(1, BILL_COLS + 1):
                    ws.cell(row=cr, column=c).fill = separator_fill
                cr += 2

                grand_taxable += proj_taxable
                grand_cgst += proj_cgst
                grand_sgst += proj_sgst
                grand_igst += proj_igst
                grand_total += proj_total

            # ── GRAND TOTAL ──
            if grand_total > 0:
                ws.cell(row=cr, column=1, value='GRAND TOTAL').font = Font(bold=True, size=12)
                ws.cell(row=cr, column=11, value=grand_taxable).font = Font(bold=True, size=12)
                ws.cell(row=cr, column=11).number_format = currency_fmt
                ws.cell(row=cr, column=12, value=grand_cgst).font = Font(bold=True, size=12)
                ws.cell(row=cr, column=12).number_format = currency_fmt
                ws.cell(row=cr, column=13, value=grand_sgst).font = Font(bold=True, size=12)
                ws.cell(row=cr, column=13).number_format = currency_fmt
                ws.cell(row=cr, column=14, value=grand_igst).font = Font(bold=True, size=12)
                ws.cell(row=cr, column=14).number_format = currency_fmt
                ws.cell(row=cr, column=15, value=grand_total).font = Font(bold=True, size=12)
                ws.cell(row=cr, column=15).number_format = currency_fmt

            # Column widths
            col_widths = {
                'A': 8, 'B': 28, 'C': 18, 'D': 18, 'E': 14,
                'F': 35, 'G': 12, 'H': 10, 'I': 8, 'J': 12,
                'K': 15, 'L': 12, 'M': 12, 'N': 12, 'O': 15
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

        # ── Fetch purchase bills and build project groups ──
        try:
            purchase_bills = db_manager.get_bills_with_line_items_for_export(
                start_date=start_date, end_date=end_date
            )
        except Exception as e:
            print(f"[!] Error fetching purchase bills for export: {e}")
            purchase_bills = []

        try:
            sales_bills = db_manager.get_sales_bills_with_line_items_for_export(
                start_date=start_date, end_date=end_date
            )
        except Exception as e:
            print(f"[!] Error fetching sales bills for export: {e}")
            sales_bills = []

        # Collect project names from bank txns, purchase bills, and sales bills
        pb_bank_projects = []
        pb_project_col = 'Project' if 'Project' in combined.columns else 'project' if not combined.empty else 'Project'
        if not combined.empty and pb_project_col in combined.columns:
            pb_bank_projects = [str(p) for p in combined[pb_project_col].dropna().unique()
                                if str(p).strip() and str(p).lower() != 'nan']
        pb_bill_projects = [str(b.get('project', '')) for b in purchase_bills
                            if str(b.get('project', '')).strip() and str(b.get('project', '')).lower() != 'nan']
        sb_bill_projects = [str(b.get('project', '')) for b in sales_bills
                            if str(b.get('project', '')).strip() and str(b.get('project', '')).lower() != 'nan']

        bills_stem_groups = build_smart_project_groups(
            pb_bank_projects, pb_bill_projects + sb_bill_projects
        )

        # Apply project filter if active
        if project:
            proj_stems = get_project_stems(project)
            bills_stem_groups = {
                s: names for s, names in bills_stem_groups.items()
                if s in proj_stems or any(normalize_project_stem(t) in proj_stems
                                          for n in names for t in str(n).split())
            }

        purchase_by_stem = match_bills_to_project_groups(purchase_bills, bills_stem_groups)
        sales_by_stem = match_bills_to_project_groups(sales_bills, bills_stem_groups)

        # TAB 7: Purchase Bills
        wb_tabs = writer.book
        if purchase_by_stem:
            ws_purchase = wb_tabs.create_sheet('Purchase Bills')
            write_bills_sheet(
                ws_purchase, purchase_by_stem, bills_stem_groups,
                sheet_title='Purchase Bills — Project Grouped',
                party_label='VENDOR', party_key='vendor_name',
                gstin_key='vendor_gstin',
                total_font=Font(name='Calibri', bold=True, color='DC2626')
            )
        else:
            pd.DataFrame({'Note': ['No purchase bills for the selected filters']}).to_excel(
                writer, sheet_name='Purchase Bills', index=False)

        # TAB 8: Sales Bills
        if sales_by_stem:
            ws_sales = wb_tabs.create_sheet('Sales Bills')
            write_bills_sheet(
                ws_sales, sales_by_stem, bills_stem_groups,
                sheet_title='Sales Bills — Project Grouped',
                party_label='BUYER', party_key='buyer_name',
                gstin_key='buyer_gstin',
                total_font=Font(name='Calibri', bold=True, color='059669')
            )
        else:
            pd.DataFrame({'Note': ['No sales bills for the selected filters']}).to_excel(
                writer, sheet_name='Sales Bills', index=False)

        # ──────────────────────────────────────────────────────────
        # TAB 9: Project Breakdown (Auditor Format)
        # ──────────────────────────────────────────────────────────
        wb = writer.book
        project_col = 'Project' if 'Project' in combined.columns else 'project' if not combined.empty else 'Project'
        v_col = 'Client/Vendor' if (not combined.empty and 'Client/Vendor' in combined.columns) else 'client_vendor'

        # Auditor-format styling (green_fill, block_bg, project_name_fill/font,
        # separator_fill already defined above for bills tabs)
        section_bold = Font(name='Calibri', bold=True, size=11)
        green_amount = Font(name='Calibri', bold=True, color='006100')
        red_amount = Font(name='Calibri', bold=True, color='DC2626')
        blue_amount = Font(name='Calibri', bold=True, color='2563EB')
        pb_currency = '#,##0.00'
        WEIGHT_UOMS = {'KGS', 'KG', 'MT', 'TONS', 'TON', 'MTS'}
        # Exclude these from Other Expense; LABOUR categories go to LABOUR PAYMENT
        EXCLUDE_CATS = {'MATERIAL PURCHASE', 'AMOUNT RECEIVED', 'SALARY AC', 'BANK CHARGES', 'DUTIES & TAX'}
        LABOUR_CATS = {'LABOUR PAYMENT', 'LABOR PAYMENT', 'LABOUR', 'LABOR'}
        NUM_COLS = 3  # A, B, C

        def pb_section_header(ws, row, text):
            """Green-filled section header spanning all columns."""
            ws.cell(row=row, column=1, value=text).font = section_bold
            for c in range(1, NUM_COLS + 1):
                ws.cell(row=row, column=c).fill = green_fill
            return row + 1

        def pb_separator(ws, row):
            """Dark separator row between project blocks."""
            for c in range(1, NUM_COLS + 1):
                ws.cell(row=row, column=c).fill = separator_fill
            return row + 1

        def pb_block_bg(ws, row):
            """Apply mild yellow background to a content row."""
            for c in range(1, NUM_COLS + 1):
                cell = ws.cell(row=row, column=c)
                if cell.fill == PatternFill(fill_type=None) or cell.fill == PatternFill():
                    cell.fill = block_bg

        try:
            export_bills = db_manager.get_bills_with_line_items_for_export(
                start_date=start_date, end_date=end_date
            )
        except Exception as e:
            print(f"[!] Error fetching bills with line items: {e}")
            export_bills = []

        # Filter export_bills by the project selection (same as combined API)
        if project:
            pb_proj_sel = parse_project_selection(project)
            if pb_proj_sel[0] or pb_proj_sel[1]:
                export_bills = [b for b in export_bills
                                if project_value_matches_selection(b.get('project'), pb_proj_sel)]

        # Collect project names from bank txns and bills
        bank_projects = []
        if not combined.empty and project_col in combined.columns:
            bank_projects = [str(p) for p in combined[project_col].dropna().unique()
                             if str(p).strip() and str(p).lower() != 'nan']
        bill_projects = [str(b.get('project', '')) for b in export_bills
                         if str(b.get('project', '')).strip() and str(b.get('project', '')).lower() != 'nan']

        stem_groups = build_smart_project_groups(bank_projects, bill_projects)

        bills_by_stem = match_bills_to_project_groups(export_bills, stem_groups)

        # Fetch labour costs from salary/attendance DB, filtered by project
        try:
            labour_costs_raw = DatabaseManager.get_labour_costs_by_project(
                start_date=start_date, end_date=end_date
            )
            if project:
                pb_proj_sel = parse_project_selection(project)
                if pb_proj_sel[0] or pb_proj_sel[1]:
                    labour_costs_raw = {k: v for k, v in labour_costs_raw.items()
                                        if project_value_matches_selection(k, pb_proj_sel)}
            labour_by_stem = match_labour_to_project_groups(labour_costs_raw, stem_groups)
        except Exception as e:
            print(f"[!] Error matching labour costs: {e}")
            labour_by_stem = {}

        if stem_groups:
            ws_pb = wb.create_sheet('Project Breakdown')
            current_row = 1

            for stem_idx, stem in enumerate(sorted(stem_groups.keys())):
                project_names = stem_groups[stem]
                group_label = stem.upper()
                group_bills = bills_by_stem.get(stem, [])
                proj_list = ', '.join(sorted(str(p) for p in project_names if str(p) != 'nan'))

                # Filter bank transactions for this project group
                if not combined.empty and project_col in combined.columns:
                    group_mask = combined[project_col].isin(project_names)
                    group_df = combined[group_mask]
                else:
                    group_df = pd.DataFrame()

                if group_df.empty and not group_bills:
                    continue

                # Track the first row of this block so we can paint background
                block_start_row = current_row

                # ════════════════════════════════════════════════════════
                # PROJECT NAME — big, bold, blue fill, white text
                # ════════════════════════════════════════════════════════
                for c in range(1, NUM_COLS + 1):
                    ws_pb.cell(row=current_row, column=c).fill = project_name_fill
                ws_pb.cell(row=current_row, column=1,
                           value=f'PROJECT :  {group_label}').font = project_name_font
                current_row += 1

                # Sub-label showing all project name variants
                ws_pb.cell(row=current_row, column=1,
                           value=f'({proj_list})').font = Font(
                    name='Calibri', italic=True, color='6B7280', size=9)
                current_row += 2  # blank row gap

                # ── OVERALL SUMMARY ──
                current_row = pb_section_header(ws_pb, current_row, 'OVERALL SUMMARY')

                # TOTAL PROJECT VALUE row (amount filled at the end)
                ws_pb.cell(row=current_row, column=1, value='TOTAL PROJECT VALUE').font = Font(bold=True)
                total_value_row = current_row
                current_row += 2  # blank row

                # ── MAIN MATERIAL PURCHASE ──
                current_row = pb_section_header(ws_pb, current_row, 'MAIN MATERIAL PURCHASE')

                # Column headers
                ws_pb.cell(row=current_row, column=2, value='WEIGHT').font = Font(bold=True)
                ws_pb.cell(row=current_row, column=3, value='AMOUNT').font = Font(bold=True)
                current_row += 1

                # Aggregate bills by vendor — use total_amount (subtotal + taxes)
                material_total = 0
                if group_bills:
                    vendor_agg = {}  # vendor_name -> {weight, amount}
                    for bill in group_bills:
                        vname = bill.get('vendor_name', 'Unknown Vendor')
                        if vname not in vendor_agg:
                            vendor_agg[vname] = {'weight': 0, 'amount': 0}
                        for item in bill.get('line_items', []):
                            uom = str(item.get('uom', '')).upper().strip()
                            if uom in WEIGHT_UOMS:
                                vendor_agg[vname]['weight'] += item.get('quantity', 0)
                        # Use bill total_amount (includes taxes) for consistent totals
                        vendor_agg[vname]['amount'] += bill.get('total_amount', 0)

                    for vname, data in sorted(vendor_agg.items(), key=lambda x: x[1]['amount'], reverse=True):
                        ws_pb.cell(row=current_row, column=1, value=vname)
                        if data['weight'] > 0:
                            ws_pb.cell(row=current_row, column=2, value=data['weight'])
                            ws_pb.cell(row=current_row, column=2).number_format = '#,##0.00'
                        ws_pb.cell(row=current_row, column=3, value=data['amount'])
                        ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                        material_total += data['amount']
                        current_row += 1

                # Material total — just the green amount (no duplicate header)
                ws_pb.cell(row=current_row, column=3, value=material_total).font = green_amount
                ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                current_row += 2  # blank row

                # ── OTHER EXPENSE ──
                current_row = pb_section_header(ws_pb, current_row, 'OTHER EXPENSE')

                other_total = 0
                if not group_df.empty:
                    expense_df = group_df[group_df['DR Amount'] > 0].copy()
                    if 'Category' in expense_df.columns:
                        # Exclude labour + standard exclusions from other expense
                        upper_cats = expense_df['Category'].str.upper().str.strip()
                        labour_mask = upper_cats.isin(LABOUR_CATS)
                        exclude_mask = expense_df['Category'].isin(EXCLUDE_CATS) | labour_mask
                        expense_df = expense_df[~exclude_mask]

                    if not expense_df.empty and v_col in expense_df.columns:
                        for (vend, cat), grp in expense_df.groupby([v_col, 'Category']):
                            amt = float(grp['DR Amount'].sum())
                            vend_str = str(vend).strip()
                            cat_str = str(cat).strip()
                            if not vend_str or vend_str.lower() in ('unknown', 'nan', '', 'unassigned'):
                                label = cat_str
                            else:
                                label = f"{cat_str} - {vend_str}"
                            ws_pb.cell(row=current_row, column=1, value=label)
                            ws_pb.cell(row=current_row, column=3, value=amt)
                            ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                            other_total += amt
                            current_row += 1
                    elif not expense_df.empty:
                        for cat, grp in expense_df.groupby('Category'):
                            amt = float(grp['DR Amount'].sum())
                            ws_pb.cell(row=current_row, column=1, value=str(cat))
                            ws_pb.cell(row=current_row, column=3, value=amt)
                            ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                            other_total += amt
                            current_row += 1

                # Other expense total (green amount only)
                ws_pb.cell(row=current_row, column=3, value=other_total).font = green_amount
                ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                current_row += 2  # blank row

                # ── LABOUR PAYMENT (blue amount, from salary/attendance DB) ──
                salary_labour = labour_by_stem.get(stem, 0)
                ws_pb.cell(row=current_row, column=1, value='LABOUR PAYMENT').font = Font(bold=True)
                if salary_labour > 0:
                    ws_pb.cell(row=current_row, column=3, value=salary_labour).font = blue_amount
                    ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                current_row += 2  # blank row

                # ── BALANCE GST PAYMENT ──
                ws_pb.cell(row=current_row, column=1, value='BALANCE GST PAYMENT').font = Font(bold=True)
                ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                current_row += 2  # blank row

                # ── OVER HEADS ──
                ws_pb.cell(row=current_row, column=1, value='OVER HEADS').font = Font(bold=True)
                ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                current_row += 2  # blank row

                # ── TOTAL EXP ──
                total_project = material_total + other_total + salary_labour
                ws_pb.cell(row=current_row, column=1, value='TOTAL EXP').font = Font(bold=True)
                ws_pb.cell(row=current_row, column=3, value=total_project).font = green_amount
                ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                current_row += 2  # blank row

                # ── BALANCE ──
                ws_pb.cell(row=current_row, column=1, value='BALANCE').font = Font(bold=True)
                ws_pb.cell(row=current_row, column=3).font = green_amount
                ws_pb.cell(row=current_row, column=3).number_format = pb_currency
                current_row += 1

                # TOTAL PROJECT VALUE left empty (user-fillable)

                block_end_row = current_row

                # Paint mild yellow background on all content rows in this block
                for r in range(block_start_row, block_end_row + 1):
                    pb_block_bg(ws_pb, r)

                # ── Dark separator + gap before next project ──
                current_row += 1
                current_row = pb_separator(ws_pb, current_row)
                current_row += 2

            # Column widths
            ws_pb.column_dimensions['A'].width = 40
            ws_pb.column_dimensions['B'].width = 15
            ws_pb.column_dimensions['C'].width = 20

        # ──────────────────────────────────────────────────────────
        # TAB 10: Labour Attendance & Salary Summary (all months in range)
        # ──────────────────────────────────────────────────────────
        import calendar as cal
        from datetime import date as date_cls

        labour_sheet_names = []
        try:
            # Use the actual date filters so all months in the range get a sheet
            labour_start = start_date
            labour_end = end_date
            if not labour_start and not labour_end:
                from datetime import date as _d
                today = _d.today()
                labour_start = f"{today.year}-{today.month:02d}-01"
                labour_end = f"{today.year}-{today.month:02d}-{cal.monthrange(today.year, today.month)[1]:02d}"

            monthly_data = DatabaseManager.get_monthly_salary_and_attendance(
                start_date=labour_start, end_date=labour_end
            )

            # Styling for labour sheets
            labour_title_font = Font(name='Calibri', bold=True, size=13, color='1A1A2E')
            labour_header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
            labour_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            summary_header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            sunday_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
            present_font = Font(color='006100')
            absent_font = Font(color='DC2626')
            labour_currency = '#,##0.00'

            for month_data in monthly_data:
                sheet_name = f"Labour {month_data['sheet_name']}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                labour_sheet_names.append(sheet_name)

                ws_l = wb.create_sheet(sheet_name)
                days_in_month = month_data['days_in_month']
                yr = month_data['year']
                mn = month_data['month_num']

                # Build attendance lookup: worker_id -> day -> {status, ot, project}
                att_map = {}
                for a in month_data['attendance']:
                    wid = a['worker_id']
                    try:
                        from dateutil import parser as dp
                        d = dp.parse(str(a['date'])).day
                    except:
                        continue
                    if wid not in att_map:
                        att_map[wid] = {}
                    att_map[wid][d] = {
                        'status': a['status'],
                        'ot': a['ot_hours'],
                        'project': a['project']
                    }

                # Filter workers by selected project(s) if active
                labour_workers = month_data['workers']
                labour_attendance = month_data['attendance']
                labour_project_breakdown = month_data['project_breakdown']
                labour_daily_headcount = month_data['daily_headcount']
                if project:
                    labour_sel = parse_project_selection(project)

                    # Step 1: Filter attendance to ONLY records on the selected project
                    project_attendance = [
                        a for a in month_data['attendance']
                        if project_value_matches_selection(a.get('project'), labour_sel)
                    ]

                    # Step 2: Get worker IDs from project-filtered attendance
                    matching_worker_ids = set(a['worker_id'] for a in project_attendance)

                    if matching_worker_ids:
                        labour_attendance = project_attendance

                        # Step 3: Recompute worker stats from project-filtered attendance only
                        from collections import defaultdict as _dd
                        worker_stats = _dd(lambda: {'present_days': 0, 'ot_hours': 0.0})
                        for a in labour_attendance:
                            wid = a['worker_id']
                            if a['status'] == 'P':
                                worker_stats[wid]['present_days'] += 1
                                worker_stats[wid]['ot_hours'] += float(a.get('ot_hours', 0) or 0)

                        labour_workers = []
                        for w in month_data['workers']:
                            if w['worker_id'] not in matching_worker_ids:
                                continue
                            wid = w['worker_id']
                            proj_days = worker_stats[wid]['present_days']
                            proj_ot = worker_stats[wid]['ot_hours']
                            base = w['base_salary_per_day']
                            proj_base_pay = round(proj_days * base, 2)
                            proj_ot_pay = round((base / 8) * proj_ot, 2) if base > 0 else 0
                            labour_workers.append({
                                **w,
                                'working_days': proj_days,
                                'ot_hours': proj_ot,
                                'base_pay': proj_base_pay,
                                'ot_pay': proj_ot_pay,
                                'total_salary': round(proj_base_pay + proj_ot_pay, 2)
                            })

                        labour_project_breakdown = [
                            p for p in month_data['project_breakdown']
                            if project_value_matches_selection(p.get('name'), labour_sel)
                        ]

                        # Rebuild att_map for filtered attendance only
                        att_map = {}
                        for a in labour_attendance:
                            wid = a['worker_id']
                            try:
                                from dateutil import parser as dp
                                d = dp.parse(str(a['date'])).day
                            except:
                                continue
                            if wid not in att_map:
                                att_map[wid] = {}
                            att_map[wid][d] = {
                                'status': a['status'],
                                'ot': a['ot_hours'],
                                'project': a['project']
                            }

                        # Recompute daily headcount from filtered attendance only
                        _dh_map = _dd(lambda: {'present': 0, 'absent': 0, 'holiday': 0, 'ot_hours': 0.0})
                        for a in labour_attendance:
                            _date_key = str(a['date'])
                            _status = a['status']
                            if _status == 'P':
                                _dh_map[_date_key]['present'] += 1
                                _dh_map[_date_key]['ot_hours'] += float(a.get('ot_hours', 0) or 0)
                            elif _status == 'A':
                                _dh_map[_date_key]['absent'] += 1
                            elif _status == 'H':
                                _dh_map[_date_key]['holiday'] += 1
                        labour_daily_headcount = [
                            {'date': d, 'present': v['present'], 'absent': v['absent'],
                             'holiday': v['holiday'], 'ot_hours': round(v['ot_hours'], 2)}
                            for d, v in sorted(_dh_map.items())
                        ]
                    else:
                        # No attendance on this project for this month — skip sheet entirely
                        continue

                # ===== ROW 1: Title =====
                ws_l.cell(row=1, column=1,
                          value=f"LABOUR ATTENDANCE FOR {month_data['sheet_name']}").font = labour_title_font
                ws_l.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

                # ===== ROW 2-3: Headers =====
                # Fixed columns
                fixed_headers = ['S.No', 'Name', 'DESIGNATION', 'TEAM']
                for ci, h in enumerate(fixed_headers, 1):
                    c = ws_l.cell(row=2, column=ci, value=h)
                    c.font = labour_header_font
                    c.fill = labour_header_fill
                    c.alignment = Alignment(horizontal='center')
                    # Row 3 empty for fixed cols
                    ws_l.cell(row=3, column=ci).fill = labour_header_fill

                # Day columns: 3 cols per day (status, OT, project)
                col = 5  # start after fixed
                day_col_starts = {}  # day -> starting column
                for day in range(1, days_in_month + 1):
                    day_col_starts[day] = col
                    dt = date_cls(yr, mn, day)
                    is_sunday = dt.weekday() == 6
                    day_label = f"{day} SUN" if is_sunday else str(day)

                    cell_h1 = ws_l.cell(row=2, column=col, value=day_label)
                    cell_h1.font = labour_header_font
                    cell_h1.fill = sunday_fill if is_sunday else labour_header_fill
                    cell_h1.alignment = Alignment(horizontal='center')
                    # Merge across 3 cols for day header
                    ws_l.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)

                    # Sub-headers
                    for si, sub in enumerate(['', 'OT', 'Pr']):
                        sc = ws_l.cell(row=3, column=col + si, value=sub)
                        sc.font = Font(size=8, bold=True)
                        sc.fill = sunday_fill if is_sunday else labour_header_fill
                        if not is_sunday:
                            sc.font = Font(size=8, bold=True, color='FFFFFF')
                        sc.alignment = Alignment(horizontal='center')
                    col += 3

                # Summary column headers
                summary_start_col = col
                sum_headers = ['TOTAL PRESENT', 'TOTAL OT', 'BASE SALARY',
                               'BASE PAY', 'OT PAY', 'TOTAL SALARY']
                # Main header merged across summary cols
                ws_l.cell(row=2, column=summary_start_col,
                          value=f"{month_data['sheet_name']} MONTH LABOUR ATTENDANCE & PAYMENT").font = Font(
                    bold=True, size=9, color='FFFFFF')
                ws_l.cell(row=2, column=summary_start_col).fill = summary_header_fill
                for c in range(summary_start_col, summary_start_col + 6):
                    ws_l.cell(row=2, column=c).fill = summary_header_fill
                ws_l.merge_cells(start_row=2, start_column=summary_start_col,
                                 end_row=2, end_column=summary_start_col + 5)
                for si, sh in enumerate(sum_headers):
                    sc = ws_l.cell(row=3, column=summary_start_col + si, value=sh)
                    sc.font = Font(bold=True, size=8, color='FFFFFF')
                    sc.fill = summary_header_fill
                    sc.alignment = Alignment(horizontal='center')

                # ===== DATA ROWS (one per worker) =====
                data_row = 4
                sno = 1
                for w in labour_workers:
                    ws_l.cell(row=data_row, column=1, value=sno)
                    ws_l.cell(row=data_row, column=2, value=w['name'])
                    ws_l.cell(row=data_row, column=3, value=w['designation'])
                    ws_l.cell(row=data_row, column=4, value=w['team'])

                    # Day-by-day columns
                    for day in range(1, days_in_month + 1):
                        dc = day_col_starts[day]
                        att = att_map.get(w['worker_id'], {}).get(day)
                        if att:
                            status_cell = ws_l.cell(row=data_row, column=dc, value=att['status'])
                            if att['status'] == 'P':
                                status_cell.font = present_font
                            elif att['status'] == 'A':
                                status_cell.font = absent_font
                            if att['ot']:
                                ws_l.cell(row=data_row, column=dc + 1, value=att['ot'])
                            if att['project']:
                                ws_l.cell(row=data_row, column=dc + 2, value=att['project'])

                    # Summary columns
                    ws_l.cell(row=data_row, column=summary_start_col, value=w['working_days'])
                    ws_l.cell(row=data_row, column=summary_start_col + 1, value=w['ot_hours'])
                    ws_l.cell(row=data_row, column=summary_start_col + 2,
                              value=w['base_salary_per_day']).number_format = labour_currency
                    ws_l.cell(row=data_row, column=summary_start_col + 3,
                              value=w['base_pay']).number_format = labour_currency
                    ws_l.cell(row=data_row, column=summary_start_col + 4,
                              value=w['ot_pay']).number_format = labour_currency
                    ws_l.cell(row=data_row, column=summary_start_col + 5,
                              value=w['total_salary']).number_format = labour_currency

                    sno += 1
                    data_row += 1

                # ===== SUMMARY SECTIONS =====
                data_row += 1  # blank row

                # Monthly Summary
                ws_l.cell(row=data_row, column=1, value='MONTHLY SUMMARY').font = Font(bold=True, size=11)
                ws_l.cell(row=data_row, column=1).fill = PatternFill(
                    start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                data_row += 1

                total_workers = len(labour_workers)
                total_present = sum(w['working_days'] for w in labour_workers)
                total_ot = sum(w['ot_hours'] for w in labour_workers)
                total_sal = sum(w['total_salary'] for w in labour_workers)

                kpi_labels = ['Total Workers', 'Total Present Days', 'Total OT Hours', 'Total Salary']
                kpi_values = [total_workers, total_present, round(total_ot, 2), total_sal]
                for ki, (kl, kv) in enumerate(zip(kpi_labels, kpi_values)):
                    c = ki * 3
                    ws_l.cell(row=data_row, column=c + 1, value=kl).font = Font(bold=True)
                    cell = ws_l.cell(row=data_row, column=c + 2, value=kv)
                    if kl == 'Total Salary':
                        cell.number_format = labour_currency
                        cell.font = Font(bold=True, color='006100')
                data_row += 2

                # Project Breakdown
                ws_l.cell(row=data_row, column=1, value='PROJECT BREAKDOWN').font = Font(bold=True, size=11)
                ws_l.cell(row=data_row, column=1).fill = PatternFill(
                    start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                data_row += 1

                pb_headers = ['Project', 'Workers', 'Working Days', 'OT Hours']
                for ci, h in enumerate(pb_headers, 1):
                    c = ws_l.cell(row=data_row, column=ci, value=h)
                    c.font = Font(bold=True, color='FFFFFF', size=9)
                    c.fill = labour_header_fill
                data_row += 1

                for proj in sorted(labour_project_breakdown, key=lambda x: x['name']):
                    ws_l.cell(row=data_row, column=1, value=proj['name'])
                    ws_l.cell(row=data_row, column=2, value=proj['workers'])
                    ws_l.cell(row=data_row, column=3, value=proj['working_days'])
                    ws_l.cell(row=data_row, column=4, value=proj['ot_hours'])
                    data_row += 1
                data_row += 1

                # Daily Headcount
                ws_l.cell(row=data_row, column=1, value='DAILY HEADCOUNT').font = Font(bold=True, size=11)
                ws_l.cell(row=data_row, column=1).fill = PatternFill(
                    start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                data_row += 1

                dh_headers = ['Day', 'Date', 'Present', 'Absent', 'Holiday', 'OT Hours']
                for ci, h in enumerate(dh_headers, 1):
                    c = ws_l.cell(row=data_row, column=ci, value=h)
                    c.font = Font(bold=True, color='FFFFFF', size=9)
                    c.fill = labour_header_fill
                data_row += 1

                day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
                for dh in labour_daily_headcount:
                    try:
                        from dateutil import parser as dp
                        dd = dp.parse(str(dh['date']))
                        ws_l.cell(row=data_row, column=1, value=day_names[dd.weekday()])
                        ws_l.cell(row=data_row, column=2, value=dd.strftime('%d/%m/%Y'))
                    except:
                        ws_l.cell(row=data_row, column=2, value=str(dh['date']))
                    ws_l.cell(row=data_row, column=3, value=dh['present'])
                    ws_l.cell(row=data_row, column=4, value=dh['absent'])
                    ws_l.cell(row=data_row, column=5, value=dh['holiday'])
                    ws_l.cell(row=data_row, column=6, value=dh['ot_hours'])
                    data_row += 1

                # Column widths
                ws_l.column_dimensions['A'].width = 5
                ws_l.column_dimensions['B'].width = 20
                ws_l.column_dimensions['C'].width = 12
                ws_l.column_dimensions['D'].width = 10
                # Day columns are narrow (3 per day)
                for day in range(1, days_in_month + 1):
                    dc = day_col_starts[day]
                    for offset, w in enumerate([3, 3, 8]):
                        col_letter = get_column_letter(dc + offset)
                        ws_l.column_dimensions[col_letter].width = w
                # Summary columns
                for si, sw in enumerate([13, 10, 12, 10, 10, 12]):
                    col_letter = get_column_letter(summary_start_col + si)
                    ws_l.column_dimensions[col_letter].width = sw

        except Exception as e:
            print(f"[!] Labour tab export error: {e}")
            import traceback
            traceback.print_exc()

        # ── Sheet ordering ──
        if len(wb.sheetnames) > 1:
            desired_order = [
                'Executive Summary', 'Expense Breakdown', 'Cashflow Analysis',
                'Vendor Breakdown'
            ]
            for bc in VALID_BANK_CODES:
                bank_config = get_bank_config(bc)
                sn = f"{bank_config['name']} Txns"
                if len(sn) > 31:
                    sn = sn[:31]
                desired_order.append(sn)
            desired_order.extend(['Project Breakdown', 'Purchase Bills', 'Sales Bills'])
            desired_order.extend(labour_sheet_names)  # Labour always last

            desired_order = [s for s in desired_order if s in wb.sheetnames]
            desired_order += [s for s in wb.sheetnames if s not in desired_order]

            wb._sheets = [wb[s] for s in desired_order]

    output.seek(0)
    filename = f"Project_Summary_{now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':

    app.run(debug=True, port=5000)
